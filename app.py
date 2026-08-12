import json
import re
import os
import base64
import hashlib
from io import BytesIO
from pathlib import Path
from datetime import date

import pandas as pd
import streamlit as st
from openai import OpenAI

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
    Image as RLImage,
)


# =========================================================
# ACCOUNTRA BRAND ASSETS
# =========================================================

BASE_DIR = Path(__file__).resolve().parent
LOGO_PATH = BASE_DIR / "assets" / "accountra_mark.png"
CREATOR_NAME = "Rohan A."
try:
    _configured_google_form_url = st.secrets.get("ACCOUNTRA_GOOGLE_FORM_URL", "")
except Exception:
    _configured_google_form_url = ""
GOOGLE_FORM_URL = str(_configured_google_form_url or os.getenv("ACCOUNTRA_GOOGLE_FORM_URL", "")).strip()


def export_widget_key(kind):
    """Create a stable, upload-specific Streamlit widget key."""
    token = str(st.session_state.get("file_token") or "no_file")
    digest = hashlib.sha256(token.encode("utf-8")).hexdigest()[:12]
    return f"{kind}_{digest}"


# =========================================================
# DISPLAY / DATA HELPERS
# =========================================================


def indian_currency(amount):
    amount = float(amount or 0)
    sign = "-" if amount < 0 else ""
    amount = abs(amount)
    integer_part = int(amount)
    decimal_part = round(amount - integer_part, 2)
    number = str(integer_part)

    if len(number) > 3:
        last_three = number[-3:]
        remaining = number[:-3]
        groups = []
        while len(remaining) > 2:
            groups.insert(0, remaining[-2:])
            remaining = remaining[:-2]
        if remaining:
            groups.insert(0, remaining)
        number = ",".join(groups) + "," + last_three

    if decimal_part:
        decimals = f"{decimal_part:.2f}"[1:]
        return f"₹{sign}{number}{decimals}"
    return f"₹{sign}{number}"


APPROVED_ASSET_HEADS = {
    "PPE",
    "Capital Work-in-Progress",
    "Intangible Assets",
    "Intangible Assets Under Development",
    "Investment Property",
    "Investments",
    "Inventories",
    "Trade Receivables",
    "Cash & Cash Equivalents",
    "Other Current Assets",
    "Other Non-current Assets",
}

APPROVED_LIABILITY_HEADS = {
    "Borrowings",
    "Current Borrowings",
    "Non-current Borrowings",
    "Trade Payables",
    "Provisions",
    "Other Current Liabilities",
    "Other Non-current Liabilities",
}

APPROVED_EQUITY_HEADS = {
    "Share Capital",
    "Other Equity",
    "Capital Account",
}

APPROVED_INCOME_HEADS = {
    "Revenue from Operations",
    "Other Income",
}

APPROVED_EXPENSE_HEADS = {
    "Cost of Materials Consumed",
    "Purchases",
    "Changes in Inventories",
    "Employee Benefits Expense",
    "Finance Costs",
    "Depreciation & Amortisation",
    "Other Expenses",
    "Tax Expense",
}

APPROVED_HEADS = (
    APPROVED_ASSET_HEADS
    | APPROVED_LIABILITY_HEADS
    | APPROVED_EQUITY_HEADS
    | APPROVED_INCOME_HEADS
    | APPROVED_EXPENSE_HEADS
)



def make_result(
    nature,
    classification,
    statement,
    reason,
    ambiguous=False,
    confidence=1.0,
    missing_information=None,
):
    return {
        "nature": nature,
        "classification": classification,
        "statement": statement,
        "ambiguous": ambiguous,
        "confidence": float(confidence),
        "reason": reason,
        "missing_information": missing_information,
    }



def clean_number_series(series):
    """Convert Excel/CSV money values to numeric safely."""
    return (
        series.astype(str)
        .str.replace("₹", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace("\\u00a0", "", regex=False)
        .str.replace("(", "-", regex=False)
        .str.replace(")", "", regex=False)
        .str.strip()
        .replace(
            {
                "": "0",
                "-": "0",
                "—": "0",
                "–": "0",
                "nan": "0",
                "None": "0",
            }
        )
        .pipe(pd.to_numeric, errors="coerce")
        .fillna(0.0)
    )



# =========================================================
# FLEXIBLE TRIAL BALANCE INGESTION
# =========================================================


def _tb_normalize_label(value):
    """Normalize workbook labels so common Excel naming variations compare cleanly."""
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _tb_find_column(header_values, aliases):
    normalized = [_tb_normalize_label(value) for value in header_values]
    alias_values = [_tb_normalize_label(alias) for alias in aliases]
    for alias in alias_values:
        if alias and alias in normalized:
            return normalized.index(alias)
    ranked = []
    for index, value in enumerate(normalized):
        if not value:
            continue
        tokens = set(value.split())
        for alias in alias_values:
            alias_tokens = set(alias.split())
            if alias_tokens and alias_tokens.issubset(tokens):
                ranked.append((len(alias_tokens), len(value), index))
    return max(ranked)[2] if ranked else None


def _tb_parse_amount(value):
    """Parse money cells that may include commas, currency symbols, Dr/Cr or brackets."""
    if value is None or pd.isna(value):
        return 0.0, None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "-"}:
        return 0.0, None
    lower = text.lower()
    side = "credit" if re.search(r"(^|\s)(cr|credit)(\s|$)", lower) else "debit" if re.search(r"(^|\s)(dr|debit)(\s|$)", lower) else None
    cleaned = re.sub(r"\b(?:dr|cr|debit|credit)\b", "", lower, flags=re.IGNORECASE)
    cleaned = cleaned.replace("₹", "").replace("$", "").replace("£", "").replace("€", "")
    number = clean_number_series(pd.Series([cleaned])).iloc[0]
    return float(number or 0), side


def _tb_candidate_from_sheet(sheet_name, frame):
    """Find a Trial Balance header row and convert the rows below it to a small table."""
    if frame is None or frame.empty:
        return None
    account_aliases = ["account", "account name", "account head", "ledger", "ledger name", "ledger account", "particulars", "description", "gl account"]
    debit_aliases = ["debit", "debits", "dr", "debit amount", "debit balance", "debit total"]
    credit_aliases = ["credit", "credits", "cr", "credit amount", "credit balance", "credit total"]
    amount_aliases = ["amount", "balance", "closing balance", "net balance", "value"]
    side_aliases = ["type", "dr cr", "debit credit", "balance type", "nature", "side"]
    candidates = []
    scan = frame.head(min(45, len(frame)))
    for header_row, row in scan.iterrows():
        header_values = row.tolist()
        account_col = _tb_find_column(header_values, account_aliases)
        debit_col = _tb_find_column(header_values, debit_aliases)
        credit_col = _tb_find_column(header_values, credit_aliases)
        amount_col = _tb_find_column(header_values, amount_aliases)
        side_col = _tb_find_column(header_values, side_aliases)
        has_split = account_col is not None and debit_col is not None and credit_col is not None
        has_combined = account_col is not None and amount_col is not None and side_col is not None
        if not has_split and not has_combined:
            continue
        score = 30 if has_split else 22
        sheet_label = _tb_normalize_label(sheet_name)
        if "trial balance" in sheet_label or sheet_label == "tb":
            score += 24
        elif "ledger" in sheet_label:
            score += 10
        elif "profit loss" in sheet_label or "balance sheet" in sheet_label or "notes" in sheet_label:
            score -= 18
        body = frame.iloc[int(header_row) + 1:].copy()
        rows = []
        for values in body.itertuples(index=False, name=None):
            if account_col >= len(values):
                continue
            raw_account = values[account_col]
            account = "" if pd.isna(raw_account) else str(raw_account).strip()
            if not account or account.lower() in {"nan", "none"}:
                continue
            debit = credit = 0.0
            if has_split:
                debit, debit_side = _tb_parse_amount(values[debit_col] if debit_col < len(values) else 0)
                credit, credit_side = _tb_parse_amount(values[credit_col] if credit_col < len(values) else 0)
                if debit_side == "credit" and credit == 0:
                    credit, debit = debit, 0.0
                if credit_side == "debit" and debit == 0:
                    debit, credit = credit, 0.0
            else:
                amount, amount_side = _tb_parse_amount(values[amount_col] if amount_col < len(values) else 0)
                side_text = str(values[side_col] if side_col < len(values) else "").lower()
                side = "credit" if "cr" in side_text or "credit" in side_text else "debit" if "dr" in side_text or "debit" in side_text else amount_side
                if side == "credit" or amount < 0:
                    credit = abs(amount)
                else:
                    debit = abs(amount)
            rows.append({"Account": account, "Debit": debit, "Credit": credit})
        total_names = {"total", "grand total", "trial balance total", "subtotal", "total trial balance", "opening balance", "closing balance"}
        rows = [row for row in rows if _tb_normalize_label(row["Account"]) not in total_names and not re.match(r"^(total|grand total|subtotal)\b", row["Account"].strip(), re.IGNORECASE)]
        usable = [row for row in rows if abs(row["Debit"]) > 0.000001 or abs(row["Credit"]) > 0.000001]
        if len(usable) < 1:
            continue
        score += min(len(usable), 20)
        candidate = pd.DataFrame(rows, columns=["Account", "Debit", "Credit"])
        confidence = min(0.99, 0.58 + (0.16 if has_split else 0.08) + (0.12 if "trial balance" in _tb_normalize_label(sheet_name) else 0) + min(len(usable), 20) / 200)
        candidates.append({"sheet": str(sheet_name), "header_row": int(header_row) + 1, "score": score, "confidence": confidence, "data": candidate, "mode": "Debit/Credit columns" if has_split else "Amount + Dr/Cr column"})
    return max(candidates, key=lambda item: item["score"]) if candidates else None


def v7_extract_trial_balance(uploaded_file):
    """Inspect an uploaded workbook and return the most likely Trial Balance table."""
    data = uploaded_file.getvalue()
    name = str(uploaded_file.name).lower()
    if name.endswith(".csv"):
        sheets = {"Uploaded CSV": pd.read_csv(BytesIO(data), header=None)}
    else:
        sheets = pd.read_excel(BytesIO(data), sheet_name=None, header=None)
    candidates = [candidate for sheet_name, frame in sheets.items() if (candidate := _tb_candidate_from_sheet(sheet_name, frame))]
    if not candidates:
        return {"data": None, "sheet": None, "header_row": None, "confidence": 0.0, "mode": None, "candidate_count": 0, "needs_review": True}
    candidates.sort(key=lambda item: item["score"], reverse=True)
    best = candidates[0]
    close_match = len(candidates) > 1 and best["score"] - candidates[1]["score"] < 10
    best["candidate_count"] = len(candidates)
    best["needs_review"] = close_match or best["confidence"] < 0.78
    return best


# =========================================================
# EXPORT / REPORTING HELPERS
# =========================================================


def make_schedule3_excel(
    company_name,
    cin,
    reporting_date,
    results_df,
    pnl_rows,
    bs_rows,
    validation_rows,
    notes_rows,
):
    """Create a professional Excel working-paper/report package."""
    wb = Workbook()

    ws_tb = wb.active
    ws_tb.title = "Classified TB"

    headers = [
        "Account",
        "Debit",
        "Credit",
        "Nature",
        "Classification",
        "Statement",
        "Ambiguous",
        "Confidence",
        "Reason",
        "Missing Information",
    ]

    ws_tb.append(headers)

    for cell in ws_tb[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    for _, row in results_df.iterrows():
        ws_tb.append([
            row.get("Account", ""),
            float(row.get("Debit", 0) or 0),
            float(row.get("Credit", 0) or 0),
            row.get("Nature", ""),
            row.get("Classification", ""),
            row.get("Statement", ""),
            bool(row.get("Ambiguous", False)),
            float(row.get("Confidence", 0) or 0),
            row.get("Reason", ""),
            row.get("Missing Information", "") or "",
        ])

    for row in ws_tb.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = '#,##0.00'

    for row in ws_tb.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = '0.00%'

    widths = {
        1: 30, 2: 16, 3: 16, 4: 14, 5: 34,
        6: 20, 7: 12, 8: 13, 9: 65, 10: 45
    }
    for col, width in widths.items():
        ws_tb.column_dimensions[get_column_letter(col)].width = width
    ws_tb.freeze_panes = "A2"
    ws_tb.auto_filter.ref = ws_tb.dimensions

    # Explicit numeric formatting for the working paper.
    for row in ws_tb.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = '₹#,##,##0.00'


    def add_report_sheet(title, columns, rows):
        ws = wb.create_sheet(title)
        ws.append([company_name])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="17365D")
        ws.append([f"CIN: {cin}" if cin else ""] + [None] * (len(columns) - 1))
        ws.append([f"Reporting date: {reporting_date.strftime('%d %B %Y')}"] + [None] * (len(columns) - 1))
        ws.append([])
        ws.append(columns)

        for cell in ws[5]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append(list(row))

        # Make numeric amounts visible as Indian Rupee values.
        if len(columns) >= 2:
            for r in range(6, ws.max_row + 1):
                cell = ws.cell(r, 2)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = '₹#,##,##0.00'
                    cell.alignment = Alignment(horizontal="right")

        # Highlight Schedule III section/total rows.
        for r in range(6, ws.max_row + 1):
            label = str(ws.cell(r, 1).value or "")
            if (label.startswith(("I.", "II.", "1.", "2.", "3."))
                    or label.startswith("Total ")):
                for c in range(1, len(columns) + 1):
                    ws.cell(r, c).font = Font(bold=True)
                    if label.startswith(("I.", "II.")):
                        ws.cell(r, c).fill = PatternFill("solid", fgColor="D9EAF7")

        widths = [62, 24, 24] if len(columns) == 2 else [34, 16, 38]
        for col, width in enumerate(widths[:len(columns)], start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A6"
        return ws

    add_report_sheet(
        "Profit & Loss",
        ["Particulars", "Amount"],
        pnl_rows,
    )

    add_report_sheet(
        "Balance Sheet",
        ["Particulars", "Amount"],
        bs_rows,
    )

    add_report_sheet(
        "Validation",
        ["Check", "Status", "Detail"],
        validation_rows,
    )

    add_report_sheet(
        "Notes",
        ["Note", "Detail"],
        notes_rows,
    )

    # Cover sheet
    cover = wb.create_sheet("Report Cover", 0)

    if LOGO_PATH.exists():
        try:
            logo = XLImage(str(LOGO_PATH))
            logo.width = 96
            logo.height = 96
            cover.add_image(logo, "A1")
        except Exception:
            pass
    cover["A12"] = company_name
    cover["A12"].font = Font(size=20, bold=True, color="FFFFFF")
    cover["A12"].fill = PatternFill("solid", fgColor="17365D")
    cover.merge_cells("A12:D13")
    cover["A15"] = "Accountra — Schedule III-style Financial Statements"
    cover["A4"].font = Font(size=14, bold=True)
    cover["A16"] = f"Reporting date: {reporting_date.strftime('%d %B %Y')}"
    cover["A17"] = f"CIN: {cin}" if cin else "CIN: Not provided"
    cover["A19"] = "Important"
    cover["A19"].font = Font(bold=True)
    cover["A20"] = (
        "This is an AI-assisted accounting preparation report. "
        "Review classifications, notes and statutory disclosures before filing."
    )
    cover["A22"] = f"Built by {CREATOR_NAME} • Accountra"
    cover["A22"].font = Font(size=9, italic=True, color="7F7F7F")
    cover.column_dimensions["A"].width = 100

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _register_pdf_fonts():
    """Register Unicode fonts so the Indian Rupee symbol renders correctly."""
    candidates = [
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
         "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf"),
        ("C:/Windows/Fonts/DejaVuSans.ttf",
         "C:/Windows/Fonts/DejaVuSans-Bold.ttf"),
        ("C:/Windows/Fonts/NotoSans-Regular.ttf",
         "C:/Windows/Fonts/NotoSans-Bold.ttf"),
    ]

    for regular, bold in candidates:
        if Path(regular).exists() and Path(bold).exists():
            try:
                pdfmetrics.registerFont(TTFont("AccountingAI-Regular", regular))
                pdfmetrics.registerFont(TTFont("AccountingAI-Bold", bold))
                return "AccountingAI-Regular", "AccountingAI-Bold"
            except Exception:
                pass

    return "Helvetica", "Helvetica-Bold"


def make_schedule3_pdf(
    company_name,
    cin,
    reporting_date,
    pnl_rows,
    bs_rows,
    validation_rows,
    notes_rows,
):
    """Create a clean PDF presentation of the generated statements."""
    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
        title=f"{company_name} - Financial Statements",
    )

    regular_font, bold_font = _register_pdf_fonts()

    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        try:
            style.fontName = regular_font
        except Exception:
            pass

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=17,
        fontName=bold_font,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontSize=10,
        fontName=regular_font,
        textColor=colors.grey,
        spaceAfter=16,
    )
    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontSize=12,
        fontName=bold_font,
        spaceBefore=12,
        spaceAfter=7,
    )

    story = []

    if LOGO_PATH.exists():
        try:
            logo = RLImage(str(LOGO_PATH), width=72, height=72)
            logo.hAlign = "CENTER"
            story.append(logo)
            story.append(Spacer(1, 5))
        except Exception:
            pass

    # Keep the PDF brand asset-free; the title below carries the Accountra identity.
    story.extend([
        Paragraph(company_name, title_style),
        Paragraph(
            f"Schedule III-style Financial Statements<br/>"
            f"As at / for the period ended {reporting_date.strftime('%d %B %Y')}"
            + (f"<br/>CIN: {cin}" if cin else ""),
            subtitle_style,
        ),
    ])

    def add_table(title, rows):
        story.append(Paragraph(title, section_style))
        data = [["Particulars", "Amount"]]
        for label, amount in rows:
            if isinstance(amount, (int, float)) and not isinstance(amount, bool):
                amount_text = indian_currency(amount)
            else:
                amount_text = str(amount or "")
            data.append([str(label), amount_text])

        table = Table(data, colWidths=[350, 150], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), bold_font),
            ("FONTNAME", (0, 1), (-1, -1), regular_font),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F7F9FC")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(table)

    add_table("Statement of Profit & Loss", pnl_rows)
    story.append(PageBreak())
    add_table("Balance Sheet", bs_rows)
    story.append(PageBreak())

    story.append(Paragraph("Final Validation", section_style))
    validation_data = [["Check", "Status", "Detail"]]
    validation_data.extend([
        [str(a), str(b), str(c)]
        for a, b, c in validation_rows
    ])
    vt = Table(validation_data, colWidths=[230, 70, 200], repeatRows=1)
    vt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(vt)

    story.append(Spacer(1, 14))
    story.append(Paragraph("Presentation Notes", section_style))
    for note, detail in notes_rows:
        story.append(Paragraph(f"<b>{note}</b>: {detail}", styles["BodyText"]))
        story.append(Spacer(1, 5))

    story.append(Spacer(1, 20))
    story.append(Paragraph(
        "AI-assisted report — review classifications and statutory disclosures "
        "before relying on the statements for statutory filing.",
        styles["Italic"],
    ))
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        f"Built by {CREATOR_NAME} • Accountra",
        ParagraphStyle(
            "CreatorCredit",
            parent=styles["Normal"],
            alignment=TA_CENTER,
            fontSize=8,
            fontName=regular_font,
            textColor=colors.grey,
        ),
    ))

    doc.build(story)
    output.seek(0)
    return output.getvalue()



# =========================================================
# DETERMINISTIC CLASSIFICATION ENGINE
# =========================================================


def classify_account(account, debit=0, credit=0):
    """
    Deterministic first-pass accounting classification.

    The order is intentional: specific payable/receivable/adjustment
    phrases are checked before broad words such as salary, capital,
    stock or loan.
    """
    name = str(account).strip().lower()
    name = re.sub(r"\s+", " ", name)

    # -----------------------------------------------------
    # YEAR-END / INVENTORY ADJUSTMENTS
    # -----------------------------------------------------
    if any(x in name for x in [
        "closing stock",
        "closing inventory",
        "closing inventories",
        "stock in trade",
        "stock-in-trade",
    ]):
        return make_result(
            "Asset",
            "Inventories",
            "Balance Sheet",
            "Closing inventory is presented as an inventory asset.",
        )

    if any(x in name for x in [
        "opening stock",
        "opening inventory",
        "opening inventories",
    ]):
        return make_result(
            "Expense",
            "Changes in Inventories",
            "Profit & Loss",
            "Opening inventory is included in determining the period's cost of goods consumed/sold.",
        )

    # -----------------------------------------------------
    # CONTRA / ADJUSTMENT TERMS THAT MUST BE REVIEWED
    # -----------------------------------------------------
    if any(x in name for x in [
        "sales return",
        "sales returns",
        "returns inward",
        "return inward",
    ]):
        return make_result(
            "Income",
            "Revenue from Operations",
            "Profit & Loss",
            "Sales returns are a reduction of operating revenue.",
        )

    if any(x in name for x in [
        "purchase return",
        "purchase returns",
        "returns outward",
        "return outward",
    ]):
        return make_result(
            "Expense",
            "Purchases",
            "Profit & Loss",
            "Purchase returns reduce purchases.",
        )

    # -----------------------------------------------------
    # SPECIFIC ASSET / LIABILITY PAIRS
    # -----------------------------------------------------
    if any(x in name for x in [
        "capital advance",
        "capital advances",
        "advance for capital goods",
        "advance for fixed asset",
        "advance against fixed asset",
    ]):
        return make_result(
            "Asset",
            "Other Non-current Assets",
            "Balance Sheet",
            "A capital advance represents an amount paid toward a capital asset and is an asset until adjusted against the asset cost.",
        )

    if any(x in name for x in [
        "advance to supplier",
        "supplier advance",
        "supplier advances",
        "advance to vendor",
        "vendor advance",
        "vendor advances",
    ]):
        return make_result(
            "Asset",
            "Other Current Assets",
            "Balance Sheet",
            "An advance paid to a supplier/vendor is an asset until the related goods or services are received.",
        )

    if any(x in name for x in [
        "advance from customer",
        "advance from customers",
        "customer advance",
        "customer advances",
        "advance received from customer",
        "income received in advance",
        "unearned income",
        "deferred revenue",
    ]):
        return make_result(
            "Liability",
            "Other Current Liabilities",
            "Balance Sheet",
            "An amount received before the related goods or services are delivered represents an obligation.",
        )

    if any(x in name for x in [
        "salary payable",
        "salaries payable",
        "wages payable",
        "employee payable",
        "bonus payable",
        "bonus payable to employees",
    ]):
        return make_result(
            "Liability",
            "Other Current Liabilities",
            "Balance Sheet",
            "Amounts payable to employees are liabilities, not current-period employee expense accounts.",
        )

    if any(x in name for x in [
        "expense payable",
        "expenses payable",
        "outstanding expense",
        "outstanding expenses",
        "accrued expense",
        "accrued expenses",
        "expense accrual",
    ]):
        return make_result(
            "Liability",
            "Other Current Liabilities",
            "Balance Sheet",
            "An outstanding/accrued expense represents an amount payable.",
        )

    if any(x in name for x in [
        "interest payable",
        "interest accrued payable",
    ]):
        return make_result(
            "Liability",
            "Other Current Liabilities",
            "Balance Sheet",
            "Interest payable is a current liability unless the facts indicate otherwise.",
        )

    if any(x in name for x in [
        "interest receivable",
        "interest accrued",
        "accrued income",
        "income accrued",
    ]):
        return make_result(
            "Asset",
            "Other Current Assets",
            "Balance Sheet",
            "Income earned but not yet received is a receivable/current asset unless facts indicate otherwise.",
        )

    # -----------------------------------------------------
    # TAX / GST / STATUTORY ASSETS AND LIABILITIES
    # -----------------------------------------------------
    if any(x in name for x in [
        "deferred tax liability",
        "deferred tax liabilities",
        "dtl",
    ]):
        return make_result(
            "Liability",
            "Other Non-current Liabilities",
            "Balance Sheet",
            "Deferred tax liabilities are treated as non-current liabilities in this model.",
        )

    if any(x in name for x in [
        "deferred tax asset",
        "deferred tax assets",
        "dta",
    ]):
        return make_result(
            "Asset",
            "Other Non-current Assets",
            "Balance Sheet",
            "A deferred tax asset is a non-current tax-related asset in this model.",
        )

    if any(x in name for x in [
        "gst receivable",
        "gst input",
        "input gst",
        "input tax credit",
        "input tax credit receivable",
        "cgst receivable",
        "sgst receivable",
        "igst receivable",
        "tds receivable",
        "tds recoverable",
        "tax receivable",
        "income tax receivable",
        "advance income tax",
        "advance tax",
    ]):
        return make_result(
            "Asset",
            "Other Current Assets",
            "Balance Sheet",
            "Tax/GST recoverable or input credits represent amounts recoverable from the tax authorities.",
        )

    if any(x in name for x in [
        "gst payable",
        "output gst",
        "output tax",
        "cgst payable",
        "sgst payable",
        "igst payable",
        "tds payable",
        "tds deducted payable",
        "tax payable",
        "professional tax payable",
        "pf payable",
        "esi payable",
        "statutory dues payable",
    ]):
        return make_result(
            "Liability",
            "Other Current Liabilities",
            "Balance Sheet",
            "Statutory taxes and dues payable represent current liabilities.",
        )

    # -----------------------------------------------------
    # PROVISIONS / PAYABLES
    # -----------------------------------------------------
    if "provision" in name:
        return make_result(
            "Liability",
            "Provisions",
            "Balance Sheet",
            "A provision represents an estimated obligation and is classified under provisions.",
        )

    if any(x in name for x in [
        "trade payable",
        "trade payables",
        "accounts payable",
        "sundry creditor",
        "sundry creditors",
        "creditor",
        "creditors",
        "supplier payable",
        "supplier payables",
    ]):
        return make_result(
            "Liability",
            "Trade Payables",
            "Balance Sheet",
            "Amounts owed to suppliers are trade payables.",
        )

    # -----------------------------------------------------
    # BORROWINGS — SPECIFIC TERMS FIRST
    # -----------------------------------------------------
    if any(x in name for x in [
        "bank overdraft",
        "bank overdrafts",
        "cash credit",
        "cash-credit",
        "working capital loan",
        "working capital borrowing",
        "short term loan",
        "short-term loan",
        "short term borrowing",
        "short-term borrowing",
        "current borrowing",
        "current borrowings",
    ]):
        return make_result(
            "Liability",
            "Current Borrowings",
            "Balance Sheet",
            "The account wording indicates a current/short-term borrowing.",
        )

    if any(x in name for x in [
        "long term loan",
        "long-term loan",
        "long term borrowing",
        "long-term borrowing",
        "non current borrowing",
        "non-current borrowing",
        "non current borrowings",
        "non-current borrowings",
    ]):
        return make_result(
            "Liability",
            "Non-current Borrowings",
            "Balance Sheet",
            "The account wording indicates a non-current/long-term borrowing.",
        )

    if any(x in name for x in [
        "loan",
        "loans",
        "borrowing",
        "borrowings",
        "term loan",
        "debenture",
        "debentures",
        "bonds payable",
        "secured loan",
        "unsecured loan",
    ]):
        return make_result(
            "Liability",
            "Borrowings",
            "Balance Sheet",
            "A loan/borrowing account represents a borrowing; current versus non-current requires maturity information.",
            ambiguous=True,
            confidence=0.90,
            missing_information="Loan maturity/repayment terms are required to determine current or non-current classification.",
        )

    # -----------------------------------------------------
    # CASH / BANK
    # -----------------------------------------------------
    if any(x in name for x in [
        "bank overdraft",
        "cash credit",
    ]):
        return make_result(
            "Liability",
            "Current Borrowings",
            "Balance Sheet",
            "Bank overdrafts/cash-credit facilities are borrowings when presented as credit balances.",
        )

    if any(x in name for x in [
        "cash",
        "petty cash",
        "cash in hand",
    ]):
        return make_result(
            "Asset",
            "Cash & Cash Equivalents",
            "Balance Sheet",
            "Cash represents cash and cash equivalents.",
        )

    if any(x in name for x in [
        "bank",
        "current account",
        "savings account",
        "fixed deposit",
        "term deposit",
        "bank deposit",
    ]):
        return make_result(
            "Asset",
            "Cash & Cash Equivalents",
            "Balance Sheet",
            "A normal debit bank/deposit balance represents cash or a cash equivalent in this model.",
        )

    # -----------------------------------------------------
    # TRADE / OTHER RECEIVABLES
    # -----------------------------------------------------
    if any(x in name for x in [
        "trade receivable",
        "trade receivables",
        "accounts receivable",
        "sundry debtor",
        "sundry debtors",
        "debtor",
        "debtors",
        "customer receivable",
        "customer receivables",
    ]):
        return make_result(
            "Asset",
            "Trade Receivables",
            "Balance Sheet",
            "Amounts due from customers are trade receivables.",
        )

    # -----------------------------------------------------
    # PREPAIDS / DEPOSITS / OTHER ASSETS
    # -----------------------------------------------------
    if any(x in name for x in [
        "prepaid",
        "prepaid expense",
        "prepaid expenses",
        "advance insurance",
        "insurance paid in advance",
    ]):
        return make_result(
            "Asset",
            "Other Current Assets",
            "Balance Sheet",
            "Prepaid expenses represent future economic benefits and are current assets in this model.",
        )

    if any(x in name for x in [
        "security deposit",
        "security deposits",
        "deposit with",
        "deposit paid",
        "refundable deposit",
    ]):
        return make_result(
            "Asset",
            "Other Non-current Assets",
            "Balance Sheet",
            "A refundable security deposit is an asset; its current/non-current presentation depends on expected recovery timing.",
            ambiguous=True,
            confidence=0.85,
            missing_information="Expected recovery period is required to determine current/non-current presentation.",
        )

    # -----------------------------------------------------
    # INVESTMENTS
    # -----------------------------------------------------
    if any(x in name for x in [
        "investment",
        "investments",
        "mutual fund",
        "mutual funds",
        "shares held",
        "bonds held",
        "securities",
    ]):
        return make_result(
            "Asset",
            "Investments",
            "Balance Sheet",
            "Investments are classified under investments in the Balance Sheet.",
        )

    # -----------------------------------------------------
    # PPE / INTANGIBLES / CAPITAL PROJECTS
    # -----------------------------------------------------
    if any(x in name for x in [
        "capital work in progress",
        "capital work-in-progress",
        "capital wip",
        "cwip",
    ]):
        return make_result(
            "Asset",
            "Capital Work-in-Progress",
            "Balance Sheet",
            "Capital expenditure on an incomplete project is classified as capital work-in-progress.",
        )

    if any(x in name for x in [
        "software under development",
        "intangible under development",
        "intangible assets under development",
        "development of software",
    ]):
        return make_result(
            "Asset",
            "Intangible Assets Under Development",
            "Balance Sheet",
            "The wording indicates an intangible asset that is still under development.",
        )

    if any(x in name for x in [
        "goodwill",
        "patent",
        "patents",
        "trademark",
        "trademarks",
        "copyright",
        "copyrights",
        "licence",
        "license",
        "franchise",
        "computer software",
        "software",
        "intangible asset",
        "intangible assets",
    ]):
        return make_result(
            "Asset",
            "Intangible Assets",
            "Balance Sheet",
            "The account represents an identifiable intangible asset.",
        )

    if any(x in name for x in [
        "land",
        "building",
        "plant",
        "machinery",
        "machine",
        "equipment",
        "furniture",
        "fixture",
        "fixtures",
        "vehicle",
        "vehicles",
        "motor car",
        "motor vehicle",
        "office equipment",
        "computer equipment",
        "computer",
        "factory building",
    ]):
        return make_result(
            "Asset",
            "PPE",
            "Balance Sheet",
            "The account represents a tangible fixed asset classified under PPE.",
        )

    if any(x in name for x in [
        "investment property",
        "investment properties",
    ]):
        return make_result(
            "Asset",
            "Investment Property",
            "Balance Sheet",
            "The account is explicitly identified as investment property.",
        )

    # -----------------------------------------------------
    # EQUITY
    # -----------------------------------------------------
    if any(x in name for x in [
        "share capital",
        "equity share capital",
        "preference share capital",
        "paid up capital",
        "paid-up capital",
    ]):
        return make_result(
            "Equity",
            "Share Capital",
            "Balance Sheet",
            "Share capital is classified under equity.",
        )

    if any(x in name for x in [
        "retained earnings",
        "retained profit",
        "surplus",
        "general reserve",
        "capital reserve",
        "securities premium",
        "share premium",
        "other reserves",
        "reserve fund",
        "reserves and surplus",
        "other equity",
    ]):
        return make_result(
            "Equity",
            "Other Equity",
            "Balance Sheet",
            "Reserves, retained earnings and similar balances are components of other equity.",
        )

    if name in {"capital", "capital account", "proprietor capital", "owner capital"}:
        return make_result(
            "Equity",
            "Capital Account",
            "Balance Sheet",
            "The capital account represents owner/proprietor equity.",
        )

    # -----------------------------------------------------
    # REVENUE / OTHER INCOME
    # -----------------------------------------------------
    if any(x in name for x in [
        "sales",
        "sale of goods",
        "sales revenue",
        "revenue from operations",
        "turnover",
        "service revenue",
        "service income",
        "consulting income",
        "consultancy income",
        "professional income",
        "operating revenue",
    ]):
        return make_result(
            "Income",
            "Revenue from Operations",
            "Profit & Loss",
            "The account represents operating revenue.",
        )

    if any(x in name for x in [
        "interest received",
        "interest income",
        "dividend income",
        "rent received",
        "rental income",
        "profit on sale of asset",
        "gain on sale of asset",
        "commission received",
        "commission income",
        "discount received",
        "other income",
        "miscellaneous income",
        "foreign exchange gain",
        "forex gain",
    ]):
        return make_result(
            "Income",
            "Other Income",
            "Profit & Loss",
            "The account represents non-operating/other income.",
        )

    # -----------------------------------------------------
    # EMPLOYEE / OPERATING EXPENSES
    # -----------------------------------------------------
    if any(x in name for x in [
        "salary",
        "salaries",
        "wages",
        "staff welfare",
        "employee benefit",
        "employee benefits",
        "bonus expense",
        "gratuity expense",
        "leave encashment expense",
        "provident fund expense",
        "pf expense",
        "esi expense",
    ]):
        return make_result(
            "Expense",
            "Employee Benefits Expense",
            "Profit & Loss",
            "Salaries, wages and employee-related costs are employee benefit expenses.",
        )

    if any(x in name for x in [
        "depreciation",
        "amortisation",
        "amortization",
    ]):
        return make_result(
            "Expense",
            "Depreciation & Amortisation",
            "Profit & Loss",
            "Depreciation/amortisation is a period expense.",
        )

    if any(x in name for x in [
        "interest expense",
        "interest paid",
        "finance cost",
        "finance costs",
        "bank charges",
        "borrowing cost",
        "borrowing costs",
        "loan interest",
    ]):
        return make_result(
            "Expense",
            "Finance Costs",
            "Profit & Loss",
            "Interest and borrowing-related charges are finance costs.",
        )

    if any(x in name for x in [
        "purchase",
        "purchases",
        "cost of goods purchased",
        "cost of materials purchased",
    ]):
        return make_result(
            "Expense",
            "Purchases",
            "Profit & Loss",
            "Purchases represent goods/materials acquired for resale or production.",
        )

    if any(x in name for x in [
        "freight inward",
        "carriage inward",
        "transport inward",
        "inward freight",
        "inward carriage",
        "direct material freight",
    ]):
        return make_result(
            "Expense",
            "Cost of Materials Consumed",
            "Profit & Loss",
            "Freight/carriage incurred to bring purchased materials into the business is treated as a material cost in this model.",
        )

    if any(x in name for x in [
        "cost of materials consumed",
        "materials consumed",
        "raw material consumed",
        "raw materials consumed",
    ]):
        return make_result(
            "Expense",
            "Cost of Materials Consumed",
            "Profit & Loss",
            "The account directly represents material consumption.",
        )

    if any(x in name for x in [
        "rent expense",
        "rent",
        "electricity",
        "power charges",
        "water charges",
        "telephone expense",
        "internet expense",
        "repairs",
        "repair expense",
        "advertising",
        "advertisement",
        "professional fees",
        "audit fees",
        "legal fees",
        "printing",
        "stationery",
        "office expenses",
        "travelling",
        "travel expense",
        "insurance expense",
        "subscription expense",
        "postage",
        "courier",
        "commission expense",
        "selling expenses",
        "miscellaneous expenses",
        "bad debts",
        "bad debt expense",
        "discount allowed",
        "foreign exchange loss",
        "forex loss",
    ]):
        return make_result(
            "Expense",
            "Other Expenses",
            "Profit & Loss",
            "The account represents an operating expense.",
        )

    if any(x in name for x in [
        "income tax expense",
        "income tax",
        "tax expense",
        "current tax",
        "current tax expense",
    ]):
        return make_result(
            "Expense",
            "Tax Expense",
            "Profit & Loss",
            "Income tax expense is presented as tax expense in Profit & Loss.",
        )

    # -----------------------------------------------------
    # UNKNOWN → AI FALLBACK
    # -----------------------------------------------------
    return None


# =========================================================
# AI FALLBACK
# =========================================================


def get_openai_client():
    """Return an OpenAI client using Streamlit secrets or environment variables."""
    api_key = None
    try:
        api_key = st.secrets.get("OPENAI_API_KEY")
    except Exception:
        pass
    api_key = api_key or os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None
    return OpenAI(api_key=api_key)


def openai_json(prompt, model=None):
    client = get_openai_client()
    if client is None:
        return None
    model = model or os.getenv("OPENAI_MODEL", "gpt-5")
    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {
                    "role": "system",
                    "content": "Return valid JSON only. Do not use markdown fences."
                },
                {"role": "user", "content": prompt},
            ],
            response_format={"type": "json_object"},
            temperature=0,
        )
        return json.loads(response.choices[0].message.content)
    except Exception as exc:
        print(f"OpenAI error: {exc}")
        return None


def classify_account_ai(account, debit, credit):
    prompt = f"""
You are an accounting classification assistant for Indian companies.
Classify this ledger account for financial statement preparation.

Account: {account}
Debit: ₹{debit}
Credit: ₹{credit}

Approved classifications only:
{sorted(APPROVED_HEADS)}

Rules:
- Never invent facts.
- If current/non-current information is missing for a borrowing or asset, mark ambiguous=true.
- Return only one approved classification.
- Confidence must be between 0 and 1.

Return exactly:
{{
  "nature": "Asset/Liability/Equity/Income/Expense",
  "classification": "approved classification",
  "statement": "Balance Sheet/Profit & Loss",
  "ambiguous": false,
  "confidence": 0.00,
  "reason": "brief explanation",
  "missing_information": null
}}
"""
    return openai_json(prompt)


def extract_pdf_text(file_bytes):
    if PdfReader is None:
        raise RuntimeError("PDF support requires the pypdf package.")
    reader = PdfReader(BytesIO(file_bytes))
    pages = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    return "\n\n".join(pages).strip()


def extract_source_text(uploaded_file):
    """Turn a source document into text/table context for the AI TB builder."""
    data = uploaded_file.getvalue()
    name = uploaded_file.name.lower()

    if name.endswith(".pdf"):
        return extract_pdf_text(data)

    if name.endswith(".csv"):
        df = pd.read_csv(BytesIO(data))
        return df.to_csv(index=False)

    if name.endswith((".xlsx", ".xls")):
        sheets = pd.read_excel(BytesIO(data), sheet_name=None)
        chunks = []
        for sheet_name, sheet_df in sheets.items():
            chunks.append(f"SHEET: {sheet_name}\n{sheet_df.fillna('').to_csv(index=False)}")
        return "\n\n".join(chunks)

    if name.endswith((".txt", ".md")):
        return data.decode("utf-8", errors="replace")

    raise ValueError("Unsupported source format.")


def build_ai_trial_balance(source_text):
    """Use AI to reconstruct an aggregated TB without inventing missing sides."""
    if not source_text or not source_text.strip():
        return None, "No readable data was found in the uploaded/pasted source."

    # Keep very large documents bounded while preserving the beginning and end.
    max_chars = 120000
    if len(source_text) > max_chars:
        source_text = (
            source_text[:90000]
            + "\n\n[...middle of source omitted for model context...]\n\n"
            + source_text[-30000:]
        )

    prompt = f"""
You are an Indian accounting data reconstruction assistant.
Your task is to convert raw accounting data into an aggregated Trial Balance for user verification.

SOURCE DATA:
{source_text}

IMPORTANT RULES:
1. Identify transactions, ledger balances, and accounting facts in the source.
2. Aggregate repeated transactions into the same account.
3. Use standard account names such as Purchases, Sales, Rent Expense, Cash, Bank, Trade Receivables, Trade Payables, etc.
4. Preserve the exact monetary amounts you can support from the source.
5. Do NOT create a balancing figure merely to make Debit equal Credit.
6. Do NOT invent cash/bank/credit settlement when the source does not support it.
7. When wording clearly supports a normal double-entry (for example, a credit purchase from a named supplier), you may infer the counterpart, but mark the entry confidence appropriately and include the evidence.
8. If a transaction cannot be completed without missing information, mark it ambiguous and explain exactly what the user needs to clarify.
9. Combine duplicate account entries after extraction.
10. Debit and credit amounts must be numeric and non-negative.
11. The final Trial Balance may be unbalanced. Never force it to balance.

Return JSON in exactly this shape:
{{
  "accounts": [
    {{
      "account": "Purchases",
      "debit": 40000,
      "credit": 0,
      "confidence": 0.95,
      "ambiguous": false,
      "evidence": "Purchased goods from ABC Traders ₹40,000",
      "missing_information": null
    }}
  ],
  "clarifications": [
    {{
      "item": "short description",
      "question": "What information is required?"
    }}
  ],
  "source_summary": "brief summary"
}}
"""

    payload = openai_json(prompt)
    if not payload or not isinstance(payload.get("accounts"), list):
        return None, "The AI could not construct a Trial Balance from the supplied data."

    rows = []
    for item in payload["accounts"]:
        if not isinstance(item, dict):
            continue
        account = str(item.get("account", "")).strip()
        if not account:
            continue
        debit = float(item.get("debit", 0) or 0)
        credit = float(item.get("credit", 0) or 0)
        if debit < 0 or credit < 0:
            continue
        rows.append({
            "Account": account,
            "Debit": debit,
            "Credit": credit,
            "Confidence": float(item.get("confidence", 0) or 0),
            "Ambiguous": bool(item.get("ambiguous", False)),
            "Evidence": str(item.get("evidence", "") or ""),
            "Missing Information": item.get("missing_information") or "",
        })

    if not rows:
        return None, "No usable accounting accounts were extracted."

    tb = pd.DataFrame(rows)
    tb["Debit"] = clean_number_series(tb["Debit"])
    tb["Credit"] = clean_number_series(tb["Credit"])

    # Aggregate repeated account names while preserving the highest-risk flags.
    tb["_key"] = tb["Account"].astype(str).str.strip().str.lower()
    grouped = []
    for _, group in tb.groupby("_key", sort=False):
        first = group.iloc[0]
        grouped.append({
            "Account": first["Account"],
            "Debit": group["Debit"].sum(),
            "Credit": group["Credit"].sum(),
            "Confidence": group["Confidence"].min(),
            "Ambiguous": bool(group["Ambiguous"].any()),
            "Evidence": " | ".join(x for x in group["Evidence"].astype(str) if x.strip())[:2000],
            "Missing Information": " | ".join(x for x in group["Missing Information"].astype(str) if x.strip())[:1000],
        })

    result = pd.DataFrame(grouped)
    return result, payload.get("clarifications", [])


# =========================================================
# COMPARATIVE / RESET HELPERS
# =========================================================


def classify_comparative_tb(comparative_df):
    """Classify a previous-year TB with the same deterministic rules used by the current TB."""
    rows = []
    for _, row in comparative_df.iterrows():
        account = str(row.get("Account", "")).strip()
        if not account:
            continue
        debit = float(row.get("Debit", 0) or 0)
        credit = float(row.get("Credit", 0) or 0)
        result = classify_account(account, debit, credit)
        if result is None:
            # For comparative information, avoid AI guessing; keep only what can be classified.
            continue
        classification = result.get("classification", "")
        if classification not in APPROVED_HEADS:
            continue
        rows.append({
            "Account": account,
            "Debit": debit,
            "Credit": credit,
            "Classification": classification,
            "Nature": result.get("nature", "Unknown"),
        })
    return pd.DataFrame(rows)


def clear_accounting_session():
    """Clear all generated/upload-derived accounting state."""
    keep_keys = {
        "input_mode", "company_name", "cin", "reporting_date", "business_nature",
        "materiality_threshold", "reset_nonce"
    }
    for key in list(st.session_state.keys()):
        if key not in keep_keys and (
            key.startswith("override_")
            or key.startswith("trial_balance_upload")
            or key.startswith("comparative_trial_balance_upload")
            or key.startswith("ai_source_upload")
            or key.startswith("generated_tb_editor")
            or key.startswith("prepare_fs_button")
            or key.startswith("confirm_ai_tb")
            or key.startswith("download_")
            or key in {
                "file_token", "prepared", "results", "comparative_results",
                "generated_tb", "generated_tb_clarifications", "generated_tb_confirmed",
                "comparative_loaded", "export_excel_bytes", "export_pdf_bytes",
                "export_signature"
            }
        ):
            del st.session_state[key]
    st.session_state["file_token"] = None
    st.session_state["prepared"] = False
    st.session_state.pop("results", None)
    st.session_state.pop("comparative_results", None)
    st.session_state.pop("comparative_loaded", None)
    st.session_state["generated_tb_confirmed"] = False
    st.session_state["reset_nonce"] = int(st.session_state.get("reset_nonce", 0)) + 1

# =========================================================
# STREAMLIT APP — SCHEDULE III PRESENTATION
# =========================================================

st.set_page_config(
    page_title="AI Accounting Software | Financial Statement Generator | Accountra",
    page_icon="assets/accountra_favicon.png",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =========================================================
# ACCOUNTRA — HIDE STREAMLIT BRANDING
# =========================================================
st.markdown(
    """
    <style>
        /* Remove Streamlit footer */
        footer {
            visibility: hidden !important;
            display: none !important;
        }

        /* Remove Streamlit top header/toolbar */
        header {
            visibility: hidden !important;
            display: none !important;
        }

        /* Remove Streamlit toolbar */
        [data-testid="stToolbar"] {
            display: none !important;
            visibility: hidden !important;
        }

        /* Remove Streamlit decoration line */
        [data-testid="stDecoration"] {
            display: none !important;
        }

        /* Remove status/developer widget */
        [data-testid="stStatusWidget"] {
            display: none !important;
        }

        /* Remove Streamlit menu */
        #MainMenu {
            display: none !important;
            visibility: hidden !important;
        }
    </style>
    """,
    unsafe_allow_html=True,
)    
st.markdown(
    """
    <style>
    /* =====================================================
       ACCOUNTRA — V2 PRODUCT UI SYSTEM
       Presentation layer only. Accounting logic is untouched.
       ===================================================== */

    :root {
        --acc-primary: #5b5cf0;
        --acc-primary-strong: #4f46e5;
        --acc-primary-soft: rgba(91,92,240,.10);
        --acc-cyan: #0891b2;
        --acc-success: #16a34a;
        --acc-warning: #d97706;
        --acc-danger: #dc2626;
        --acc-ink: #111827;
        --acc-muted: #667085;
        --acc-line: #e7eaf0;
        --acc-line-strong: #d9dee8;
        --acc-bg: #f7f8fc;
        --acc-surface: #ffffff;
        --acc-surface-soft: #f9fafc;
        --acc-sidebar: #fbfbfd;
        --acc-shadow-sm: 0 1px 2px rgba(16,24,40,.04), 0 5px 18px rgba(16,24,40,.045);
        --acc-shadow-md: 0 12px 32px rgba(16,24,40,.08);
        --acc-radius: 16px;
        --acc-radius-lg: 22px;
    }

    .stApp,
    [data-testid="stAppViewContainer"],
    [data-testid="stMain"] {
        background: var(--acc-bg) !important;
        color: var(--acc-ink) !important;
    }

    .block-container {
        max-width: 1440px !important;
        padding-top: 1.15rem !important;
        padding-bottom: 4rem !important;
        padding-left: clamp(1rem, 3vw, 2.4rem) !important;
        padding-right: clamp(1rem, 3vw, 2.4rem) !important;
    }

    html, body, [class*="css"] {
        font-family: Inter, ui-sans-serif, system-ui, -apple-system,
                     BlinkMacSystemFont, "Segoe UI", sans-serif;
    }

    h1, h2, h3, h4, p, label, span, div {
        text-rendering: optimizeLegibility;
    }

    h1, h2, h3, h4 {
        letter-spacing: -.025em;
        color: var(--acc-ink);
    }

    h2 { margin-top: 1.45rem !important; margin-bottom: .65rem !important; }
    h3 { margin-top: 1rem !important; }

    /* ---------- Product brand ---------- */
    .acc-brand {
        display: inline-flex;
        align-items: center;
        gap: .72rem;
        text-decoration: none;
        color: var(--acc-ink);
    }
    .acc-logo-mark {
        width: 38px;
        height: 38px;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        border-radius: 11px;
        color: #fff;
        font-weight: 900;
        font-size: 1.05rem;
        letter-spacing: -.06em;
        background: linear-gradient(145deg, #6366f1, #4f46e5 58%, #0891b2);
        box-shadow: 0 8px 18px rgba(79,70,229,.22);
    }
    .acc-brand-name {
        font-size: 1.14rem;
        line-height: 1;
        font-weight: 850;
        letter-spacing: -.035em;
    }
    .acc-brand-sub {
        display: block;
        margin-top: .23rem;
        font-size: .70rem;
        font-weight: 650;
        color: var(--acc-muted);
        letter-spacing: .02em;
    }

    .workspace-topbar {
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 1rem;
        padding: .65rem .8rem .9rem .8rem;
        margin-bottom: .6rem;
        border-bottom: 1px solid var(--acc-line);
    }
    .workspace-context {
        text-align: right;
        color: var(--acc-muted);
        font-size: .78rem;
        line-height: 1.45;
    }
    .workspace-context strong { color: var(--acc-ink); font-weight: 750; }

    /* ---------- Buttons ---------- */
    .stButton > button,
    .stFormSubmitButton > button,
    [data-testid="stDownloadButton"] button {
        min-height: 2.7rem !important;
        border-radius: 11px !important;
        border: 1px solid var(--acc-line-strong) !important;
        background: var(--acc-surface) !important;
        color: var(--acc-ink) !important;
        font-weight: 720 !important;
        letter-spacing: -.01em;
        box-shadow: 0 1px 2px rgba(16,24,40,.035);
        transition: transform .16s ease, box-shadow .16s ease,
                    border-color .16s ease, background .16s ease;
    }
    .stButton > button:hover,
    .stFormSubmitButton > button:hover,
    [data-testid="stDownloadButton"] button:hover {
        transform: translateY(-1px);
        border-color: rgba(91,92,240,.38) !important;
        box-shadow: 0 7px 18px rgba(16,24,40,.09);
    }
    .stButton > button[kind="primary"],
    .stFormSubmitButton > button[kind="primary"] {
        background: linear-gradient(135deg, #5b5cf0, #4f46e5) !important;
        color: #fff !important;
        border-color: transparent !important;
        box-shadow: 0 8px 18px rgba(79,70,229,.20);
    }
    .stButton > button[kind="primary"]:hover,
    .stFormSubmitButton > button[kind="primary"]:hover {
        box-shadow: 0 12px 24px rgba(79,70,229,.28);
    }

    /* ---------- Inputs ---------- */
    div[data-baseweb="input"] > div,
    div[data-baseweb="textarea"] > div,
    div[data-baseweb="select"] > div,
    div[data-baseweb="popover"] {
        border-radius: 11px !important;
    }
    div[data-baseweb="input"] > div,
    div[data-baseweb="textarea"] > div,
    div[data-baseweb="select"] > div {
        background: var(--acc-surface) !important;
        border-color: var(--acc-line-strong) !important;
        color: var(--acc-ink) !important;
        transition: border-color .16s ease, box-shadow .16s ease;
    }
    div[data-baseweb="input"] > div:focus-within,
    div[data-baseweb="textarea"] > div:focus-within,
    div[data-baseweb="select"] > div:focus-within {
        border-color: rgba(91,92,240,.58) !important;
        box-shadow: 0 0 0 3px rgba(91,92,240,.11) !important;
    }
    input, textarea { color: var(--acc-ink) !important; }

    /* ---------- Sidebar ---------- */
    section[data-testid="stSidebar"] {
        background: var(--acc-sidebar) !important;
        border-right: 1px solid var(--acc-line) !important;
        box-shadow: 5px 0 24px rgba(16,24,40,.025);
    }
    section[data-testid="stSidebar"] > div {
        padding-top: .85rem !important;
        padding-left: .85rem !important;
        padding-right: .85rem !important;
    }
    section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {
        color: var(--acc-muted);
    }
    .sidebar-brand {
        padding: .2rem .25rem .9rem;
        margin-bottom: .4rem;
        border-bottom: 1px solid var(--acc-line);
    }
    .sidebar-section-label {
        margin: 1.1rem .3rem .45rem;
        color: var(--acc-muted);
        font-size: .68rem;
        font-weight: 800;
        letter-spacing: .08em;
        text-transform: uppercase;
    }

    /* ---------- Cards ---------- */
    .section-card {
        position: relative;
        height: 100%;
        padding: 1.15rem 1.2rem;
        border: 1px solid var(--acc-line);
        border-radius: var(--acc-radius);
        margin: .45rem 0 1rem;
        background: var(--acc-surface);
        box-shadow: var(--acc-shadow-sm);
        transition: transform .18s ease, box-shadow .18s ease, border-color .18s ease;
    }
    .section-card:hover {
        transform: translateY(-2px);
        border-color: rgba(91,92,240,.24);
        box-shadow: var(--acc-shadow-md);
    }
    .fs-title { font-size: 1.02rem; font-weight: 800; letter-spacing: -.018em; color: var(--acc-ink); }
    .fs-subtitle { margin-top: .32rem; font-size: .86rem; line-height: 1.55; color: var(--acc-muted); }

    .app-hero {
        position: relative;
        overflow: hidden;
        padding: 1.5rem 1.65rem;
        border: 1px solid rgba(91,92,240,.18);
        border-radius: var(--acc-radius-lg);
        background: linear-gradient(135deg, rgba(91,92,240,.11), rgba(8,145,178,.055));
        box-shadow: var(--acc-shadow-sm);
        margin-bottom: 1.2rem;
    }
    .app-hero h1 { margin: 0 0 .3rem; font-size: clamp(1.7rem, 3vw, 2.25rem); font-weight: 850; }
    .app-hero p { margin: 0; color: var(--acc-muted); line-height: 1.6; }

    /* ---------- File uploader ---------- */
    [data-testid="stFileUploaderDropzone"] {
        min-height: 155px;
        border: 1.5px dashed rgba(91,92,240,.34) !important;
        border-radius: 15px !important;
        background: var(--acc-surface-soft) !important;
        transition: border-color .18s ease, background .18s ease, transform .18s ease;
    }
    [data-testid="stFileUploaderDropzone"]:hover {
        border-color: rgba(91,92,240,.65) !important;
        background: var(--acc-primary-soft) !important;
        transform: translateY(-1px);
    }

    /* ---------- Metrics ---------- */
    [data-testid="stMetric"] {
        border: 1px solid var(--acc-line);
        border-radius: 14px;
        padding: .82rem .95rem;
        background: var(--acc-surface);
        box-shadow: var(--acc-shadow-sm);
    }
    [data-testid="stMetricLabel"] { color: var(--acc-muted) !important; font-size: .72rem !important; font-weight: 750 !important; text-transform: uppercase; letter-spacing: .045em; }
    [data-testid="stMetricValue"] { color: var(--acc-ink) !important; font-weight: 850 !important; letter-spacing: -.035em; }

    /* ---------- Tables ---------- */
    [data-testid="stDataFrame"] {
        border: 1px solid var(--acc-line) !important;
        border-radius: 14px !important;
        overflow: hidden;
        background: var(--acc-surface) !important;
        box-shadow: var(--acc-shadow-sm);
    }
    [data-testid="stDataFrame"] > div { border-radius: 14px !important; }

    /* ---------- Tabs / expanders / alerts ---------- */
    button[data-baseweb="tab"] {
        color: var(--acc-muted) !important;
        font-weight: 720 !important;
        border-radius: 9px 9px 0 0 !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] { color: var(--acc-primary) !important; }
    [data-testid="stExpander"] {
        border: 1px solid var(--acc-line) !important;
        border-radius: 13px !important;
        overflow: hidden;
        background: var(--acc-surface) !important;
        box-shadow: 0 3px 14px rgba(16,24,40,.035);
    }
    div[data-testid="stAlert"] {
        border-radius: 12px !important;
        border-width: 1px !important;
        box-shadow: none !important;
    }

    /* ---------- Statement tables ---------- */
    .statement-card {
        border: 1px solid var(--acc-line);
        border-radius: 16px;
        overflow: hidden;
        margin: .75rem 0 1.4rem;
        background: var(--acc-surface);
        box-shadow: var(--acc-shadow-sm);
    }
    .statement-head {
        padding: 1rem 1.15rem .85rem;
        border-bottom: 1px solid var(--acc-line);
        background: linear-gradient(135deg, rgba(91,92,240,.075), rgba(8,145,178,.025));
    }
    .statement-title { font-size: 1.15rem; font-weight: 850; color: var(--acc-ink); }
    .statement-subtitle { margin-top: .2rem; font-size: .82rem; color: var(--acc-muted); }
    .statement-scroll { overflow-x: auto; }
    .statement-table { width: 100%; border-collapse: collapse; min-width: 720px; color: var(--acc-ink); }
    .statement-table th { padding: .75rem .85rem; border-bottom: 1px solid var(--acc-line-strong); font-size: .70rem; text-transform: uppercase; letter-spacing: .055em; color: var(--acc-muted); text-align: left; white-space: nowrap; }
    .statement-table td { padding: .65rem .85rem; border-bottom: 1px solid var(--acc-line); font-size: .88rem; vertical-align: middle; }
    .statement-table tbody tr:hover { background: var(--acc-primary-soft); }
    .statement-table .note { width: 85px; text-align: center; }
    .statement-table .amount { width: 175px; text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums; }
    .statement-table .section td { font-weight: 850; background: var(--acc-primary-soft); border-top: 1px solid rgba(91,92,240,.18); }
    .statement-table .subsection td { font-weight: 760; }
    .statement-table .indent .particular { padding-left: 1.7rem; }
    .statement-table .total td { font-weight: 800; border-top: 1px solid var(--acc-line-strong); }
    .statement-table .subtotal td { font-weight: 800; border-top: 1px dashed var(--acc-line-strong); }
    .statement-table .grand-total td { font-weight: 900; border-top: 2px solid rgba(91,92,240,.34); border-bottom: 2px double rgba(91,92,240,.34); background: var(--acc-primary-soft); }

    /* ---------- Export links ---------- */
    .acc-download-link {
        display:block;
        width:100%;
        box-sizing:border-box;
        padding:.72rem 1rem;
        border-radius:11px;
        text-align:center;
        text-decoration:none !important;
        font-weight:760;
        color:#fff !important;
        background:linear-gradient(135deg,#5b5cf0,#4f46e5);
        border:1px solid transparent;
        box-shadow:0 7px 18px rgba(79,70,229,.18);
        transition:transform .16s ease,box-shadow .16s ease;
    }
    .acc-download-link:hover {
        transform:translateY(-1px);
        box-shadow:0 11px 24px rgba(79,70,229,.26);
    }
    /* ---------- Validation ---------- */
    .validation-pass { padding: .72rem .85rem; border-radius: 11px; border: 1px solid rgba(34,197,94,.22); background: rgba(34,197,94,.055); margin: .35rem 0; }
    .validation-grid { display: grid; grid-template-columns: repeat(2,minmax(0,1fr)); gap: .55rem; }
    .validation-item { padding: .72rem .82rem; border: 1px solid var(--acc-line); border-radius: 11px; background: var(--acc-surface); font-size: .86rem; }
    .validation-item.pass { border-color: rgba(34,197,94,.25); background: rgba(34,197,94,.055); }
    .validation-item.review { border-color: rgba(245,158,11,.28); background: rgba(245,158,11,.055); }

    /* ---------- Landing page ---------- */
    .landing-shell { max-width: 1180px; margin: .8rem auto 0; }
    .landing-brandbar { display:flex; align-items:center; justify-content:space-between; gap:1rem; margin: .35rem 0 1.2rem; }
    .landing-pill { display:inline-flex; align-items:center; gap:.4rem; padding:.38rem .65rem; border:1px solid var(--acc-line); border-radius:999px; background:var(--acc-surface); color:var(--acc-muted); font-size:.72rem; font-weight:750; }
    .landing-hero { position:relative; overflow:hidden; padding:3.3rem 3.1rem; border:1px solid rgba(91,92,240,.17); border-radius:28px; background: radial-gradient(circle at 88% 10%, rgba(34,211,238,.13), transparent 28%), radial-gradient(circle at 8% 92%, rgba(99,102,241,.13), transparent 34%), var(--acc-surface); box-shadow: var(--acc-shadow-md); }
    .landing-hero::after { content:""; position:absolute; width:360px; height:360px; right:-190px; top:-190px; border-radius:50%; border:1px solid rgba(91,92,240,.14); box-shadow:0 0 0 24px rgba(91,92,240,.025),0 0 0 48px rgba(91,92,240,.015); pointer-events:none; }
    .landing-eyebrow { display:inline-flex; align-items:center; gap:.45rem; padding:.42rem .72rem; border:1px solid rgba(91,92,240,.18); border-radius:999px; background:var(--acc-primary-soft); color:var(--acc-primary); font-size:.72rem; font-weight:820; letter-spacing:.045em; }
    .landing-title { max-width:860px; margin-top:1rem; font-size:clamp(2.55rem,5.2vw,4.55rem); line-height:1.02; font-weight:900; letter-spacing:-.06em; color:var(--acc-ink); }
    .landing-title span { background:linear-gradient(100deg,#6366f1 5%,#4f46e5 52%,#0891b2 95%); -webkit-background-clip:text; background-clip:text; color:transparent; }
    .landing-copy { max-width:720px; margin-top:1.1rem; font-size:1.05rem; line-height:1.72; color:var(--acc-muted); }
    .landing-visual { min-height:340px; display:flex; align-items:center; justify-content:center; animation:acc-float 5s ease-in-out infinite; }
    .landing-feature { height:100%; min-height:108px; padding:1.05rem 1.1rem; border:1px solid var(--acc-line); border-radius:15px; background:var(--acc-surface); box-shadow:var(--acc-shadow-sm); transition:transform .18s ease,box-shadow .18s ease,border-color .18s ease; }
    .landing-feature:hover { transform:translateY(-3px); border-color:rgba(91,92,240,.25); box-shadow:var(--acc-shadow-md); }
    .landing-feature-title { font-weight:820; font-size:.96rem; margin-bottom:.35rem; color:var(--acc-ink); }
    .landing-feature-text { font-size:.83rem; line-height:1.55; color:var(--acc-muted); }
    .landing-note { text-align:center; margin-top:1.1rem; font-size:.76rem; color:var(--acc-muted); }

    /* ---------- Feedback ---------- */
    .feedback-card { padding:1.1rem 1.15rem; border:1px solid var(--acc-line); border-radius:15px; background:var(--acc-surface); box-shadow:var(--acc-shadow-sm); }

    hr { margin:1.35rem 0 !important; border:0 !important; border-top:1px solid var(--acc-line) !important; }

    @keyframes acc-float { 0%,100% { transform:translateY(0); } 50% { transform:translateY(-5px); } }
    @media (prefers-reduced-motion: reduce) { *,*::before,*::after { animation:none !important; transition:none !important; } }

    @media (max-width: 900px) {
        .block-container { padding-left: .95rem !important; padding-right: .95rem !important; }
        .landing-hero { padding:2.35rem 1.45rem; border-radius:22px; }
        .landing-title { font-size:clamp(2.35rem,10vw,3.6rem); }
        .workspace-topbar { align-items:flex-start; }
        .workspace-context { display:none; }
    }
    @media (max-width: 640px) {
        .block-container { padding-top:.65rem !important; }
        .landing-shell { margin-top:.35rem; }
        .landing-hero { padding:1.8rem 1.05rem; }
        .landing-copy { font-size:.96rem; }
        .landing-visual { min-height:250px; }
        .statement-table { min-width:650px; }
        .validation-grid { grid-template-columns:1fr; }
        .acc-brand-sub { display:none; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# =========================================================
# SEO LANDING PAGE — TRIAL BALANCE TO FINANCIAL STATEMENTS
# =========================================================

if st.query_params.get("seo") == "trial-balance-to-financial-statements":
    st.markdown(
        """
        <div class="landing-shell">
            <div class="landing-hero">
                <div class="landing-eyebrow">📑 TRIAL BALANCE TO FINANCIAL STATEMENTS</div>
                <div class="landing-title">Trial Balance to Financial Statements</div>
                <div class="landing-subtitle">
                    Turn a structured Trial Balance into Schedule III-style Profit &amp; Loss
                    and Balance Sheet statements with AI-assisted account classification,
                    review flags, and validation checks.
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    seo_col1, seo_col2 = st.columns([1.15, 0.85], gap="large")
    with seo_col1:
        st.markdown("## How Accountra converts a Trial Balance into financial statements")
        st.write(
            """Accountra starts with your Excel or CSV Trial Balance, checks that the
            debit and credit totals are balanced, classifies accounts into financial
            statement heads, lets you review ambiguous classifications, and prepares
            Profit & Loss and Balance Sheet statements."""
        )

        st.markdown("### The workflow")
        st.markdown(
            """
            **1. Upload your Trial Balance** — Start with an Excel or CSV file.\n\n
            **2. Validate the Trial Balance** — Check the debit and credit totals before preparation.\n\n
            **3. Classify accounts** — Use AI-assisted classification with confidence and review flags.\n\n
            **4. Review ambiguous accounts** — Correct classifications when additional context is needed.\n\n
            **5. Generate financial statements** — Prepare Profit &amp; Loss and Balance Sheet statements in a Schedule III-style presentation.\n\n
            **6. Validate the output** — Review balances, missing information, and unusual classifications before using the results.
            """
        )

    with seo_col2:
        st.markdown("### What you can generate")
        st.markdown(
            """
            - Profit &amp; Loss statement
            - Balance Sheet
            - Schedule III-style presentation
            - Comparative period information when supplied
            - Account-level notes and review insights
            - Excel working-paper export
            - PDF presentation export
            """
        )
        st.info(
            """Accountra is an accounting workflow assistant. Always review generated
            classifications and statements before statutory or professional use."""
        )

    st.divider()
    st.markdown("## Frequently asked questions")

    with st.expander("Can Accountra generate financial statements from a Trial Balance?"):
        st.write(
            """Yes. Accountra is designed to take a Trial Balance, classify accounts,
            and prepare Profit & Loss and Balance Sheet statements in a Schedule III-style layout."""
        )

    with st.expander("What Trial Balance files can I upload?"):
        st.write(
            """The workflow supports Excel and CSV Trial Balance files. The uploaded
            data is then validated and prepared for account classification."""
        )

    with st.expander("Does Accountra automatically classify every account?"):
        st.write(
            """Accountra provides AI-assisted account classification and identifies
            items that may need review. You should review ambiguous or low-confidence items."""
        )

    with st.expander("Can I review the generated financial statements?"):
        st.write(
            """Yes. The workflow includes review and validation steps before the final
            financial statement exports are generated."""
        )

    st.markdown("## Ready to turn your Trial Balance into financial statements?")
    if st.button("🚀 Start with a Trial Balance", type="primary", use_container_width=True, key="seo_tb_start"):
        st.query_params.clear()
        st.session_state["app_page"] = "workspace"
        st.rerun()

    st.markdown(
        '<div class="landing-note"><a href="/">← Back to Accountra</a></div>',
        unsafe_allow_html=True,
    )
    st.stop()



# =========================================================
# ACCOUNTRA V7 — PRODUCT WORKSPACE UI
# =========================================================

# V7 keeps the original accounting/classification/reporting engine intact.
# Only the presentation/workflow layer below is rebuilt.

if "app_page" not in st.session_state:
    st.session_state["app_page"] = "dashboard"
if "workspace_section" not in st.session_state:
    st.session_state["workspace_section"] = "upload"
if "company_name" not in st.session_state:
    st.session_state["company_name"] = "ABC Private Limited"
if "cin" not in st.session_state:
    st.session_state["cin"] = ""
if "reporting_date" not in st.session_state:
    st.session_state["reporting_date"] = date.today()
if "materiality_threshold" not in st.session_state:
    st.session_state["materiality_threshold"] = 20.0
if "prepared" not in st.session_state:
    st.session_state["prepared"] = False
if "reset_nonce" not in st.session_state:
    st.session_state["reset_nonce"] = 0

st.markdown(f"""
<style>
:root {{
  --a-primary:#5865f2; --a-primary-2:#3b82f6; --a-cyan:#06b6d4;
  --a-ink:#111827;
  --a-muted:#667085;
  --a-bg:#f5f7fb;
  --a-surface:#ffffff;
  --a-surface-2:#f8fafc;
  --a-line:#e5eaf2; --a-line-strong:#d6ddea; --a-primary-soft:rgba(88,101,242,.08);
  --a-success:#16a34a; --a-warning:#d97706; --a-danger:#dc2626;
  --a-shadow:0 14px 45px rgba(15,23,42,.08);
  --a-radius:20px;
}}
.stApp {{ background:var(--a-bg)!important; color:var(--a-ink)!important; }}
.main .block-container {{ max-width:1450px!important; padding:1.6rem 3.2rem 4rem!important; }}
body, .stApp, .stMarkdown, p, label, input, textarea, button {{ font-size:1rem!important; }}
[data-testid="stHeader"], header, footer, #MainMenu, [data-testid="stToolbar"] {{ display:none!important; }}
[data-testid="stSidebar"] {{ display:none!important; }}

.a-topbar {{ display:flex; align-items:center; justify-content:space-between; gap:1rem; padding:.2rem 0 1rem; border-bottom:1px solid var(--a-line); margin-bottom:.35rem; }}
.site-nav {{ display:flex; justify-content:flex-end; gap:.45rem; margin:0 0 1rem; }}
.site-nav-note {{ margin-right:auto; align-self:center; color:var(--a-muted); font-size:.76rem; font-weight:750; }}
.site-nav .stButton button {{ min-height:38px!important; padding:.35rem .85rem!important; border:1px solid var(--a-line)!important; background:var(--a-surface)!important; color:var(--a-ink)!important; font-size:.82rem!important; box-shadow:none!important; }}
.site-nav .stButton button:hover {{ border-color:rgba(88,101,242,.35)!important; color:var(--a-primary)!important; box-shadow:0 5px 16px rgba(15,23,42,.07)!important; }}
.a-brand {{ display:flex; align-items:center; gap:.75rem; font-weight:800; color:var(--a-ink); }}
.a-logo {{ width:40px; height:40px; border-radius:12px; display:grid; place-items:center; background:linear-gradient(135deg,var(--a-primary),var(--a-cyan)); color:white; font-weight:900; font-size:1.25rem; box-shadow:0 8px 22px rgba(88,101,242,.25); }}
.a-brand-name {{ font-size:1.18rem; letter-spacing:-.02em; }}
.a-brand-sub {{ display:block; color:var(--a-muted); font-size:.78rem; font-weight:600; margin-top:.1rem; }}
.a-top-actions {{ display:flex; align-items:center; justify-content:flex-end; gap:.8rem; }}
.a-top-meta {{ color:var(--a-muted); font-size:.86rem; font-weight:700; }}
.a-session-pill {{ display:inline-flex; align-items:center; gap:.45rem; padding:.45rem .7rem; border:1px solid var(--a-line); border-radius:999px; background:var(--a-surface); color:var(--a-muted); font-size:.76rem; font-weight:800; white-space:nowrap; }}
.a-session-dot {{ width:.48rem; height:.48rem; border-radius:50%; background:var(--a-success); box-shadow:0 0 0 3px rgba(22,163,74,.12); }}

.hero {{ position:relative; overflow:hidden; border:1px solid rgba(88,101,242,.18); border-radius:28px; padding:4.1rem 4.5rem 3.35rem; background:linear-gradient(135deg,#ffffff 0%,#eef2ff 62%,#e0f7ff 100%); box-shadow:var(--a-shadow); animation:accountra-rise .55s ease both; }}
.hero:after {{ content:""; position:absolute; width:380px; height:380px; border-radius:50%; right:-130px; top:-210px; border:1px solid rgba(88,101,242,.18); box-shadow:0 0 0 55px rgba(88,101,242,.03),0 0 0 110px rgba(88,101,242,.02); }}
.eyebrow {{ color:var(--a-primary); font-weight:900; font-size:.82rem; letter-spacing:.14em; text-transform:uppercase; margin-bottom:1rem; }}
.hero h1 {{ font-size:clamp(3rem,6vw,5.7rem)!important; line-height:.98!important; letter-spacing:-.055em!important; margin:0!important; max-width:1050px; color:var(--a-ink)!important; }}
.hero h1 span {{ background:linear-gradient(90deg,var(--a-primary),var(--a-cyan)); -webkit-background-clip:text; background-clip:text; color:transparent; }}
.hero-copy {{ max-width:680px; margin-top:1.2rem; font-size:1.25rem; line-height:1.65; color:var(--a-muted); }}
.hero-actions {{ margin-top:1.35rem; }}
.hero-actions-copy {{ color:var(--a-ink); font-weight:800; font-size:1rem; }}
.hero-actions-note {{ color:var(--a-muted); font-size:.86rem; line-height:1.5; margin-top:.35rem; }}
.hero-proof {{ display:flex; flex-wrap:wrap; gap:.55rem 1rem; margin-top:.8rem; color:var(--a-muted); font-size:.78rem; font-weight:700; }}
.hero-proof span {{ display:inline-flex; align-items:center; gap:.35rem; }}
.hero-proof span::before {{ content:"✓"; color:var(--a-success); font-weight:900; }}
.dashboard-rail {{ display:grid; grid-template-columns:repeat(3,1fr); gap:.8rem; margin:1rem 0 1.25rem; }}
.dashboard-card {{ position:relative; overflow:hidden; min-height:112px; padding:1.05rem 1.1rem; border:1px solid var(--a-line); border-radius:17px; background:var(--a-surface); box-shadow:0 5px 20px rgba(15,23,42,.035); animation:accountra-rise .55s ease both; }}
.dashboard-card:nth-child(2) {{ animation-delay:.08s; }}
.dashboard-card:nth-child(3) {{ animation-delay:.16s; }}
.dashboard-card::after {{ content:""; position:absolute; right:-24px; bottom:-35px; width:100px; height:100px; border:1px solid rgba(88,101,242,.1); border-radius:50%; box-shadow:0 0 0 18px rgba(88,101,242,.025); }}
.dashboard-card-num {{ color:var(--a-primary); font-size:.72rem; font-weight:900; letter-spacing:.09em; }}
.dashboard-card-title {{ margin-top:.55rem; color:var(--a-ink); font-size:1.12rem; font-weight:900; }}
.dashboard-card-copy {{ margin-top:.25rem; color:var(--a-muted); font-size:.86rem; line-height:1.45; }}
.dashboard-strip {{ display:flex; align-items:center; justify-content:space-between; gap:1rem; margin-top:1.35rem; padding:1rem 1.1rem; border:1px solid var(--a-line); border-radius:16px; background:var(--a-surface); }}
.dashboard-strip strong {{ color:var(--a-ink); font-size:1rem; }}
.dashboard-strip span {{ color:var(--a-muted); font-size:.84rem; line-height:1.45; }}
.page-hero {{ padding:1.6rem 1.7rem; margin-bottom:1rem; border:1px solid rgba(88,101,242,.16); border-radius:20px; background:linear-gradient(135deg,rgba(88,101,242,.08),rgba(6,182,212,.06)); animation:accountra-rise .5s ease both; }}
.page-eyebrow {{ color:var(--a-primary); font-size:.72rem; font-weight:900; letter-spacing:.1em; text-transform:uppercase; }}
.page-title {{ margin-top:.4rem; color:var(--a-ink); font-size:clamp(2rem,4vw,3.3rem); font-weight:900; letter-spacing:-.05em; }}
.page-copy {{ max-width:720px; margin-top:.45rem; color:var(--a-muted); font-size:1.05rem; line-height:1.6; }}
.about-grid {{ display:grid; grid-template-columns:1.1fr .9fr; gap:1rem; margin-top:1rem; }}
.about-card, .contact-card {{ padding:1.3rem; border:1px solid var(--a-line); border-radius:18px; background:var(--a-surface); box-shadow:0 5px 20px rgba(15,23,42,.035); }}
.about-card strong, .contact-card strong {{ display:block; color:var(--a-ink); font-size:1.15rem; }}
.about-card p, .contact-card p {{ color:var(--a-muted); line-height:1.6; margin:.45rem 0 0; }}
.about-points {{ display:grid; gap:.65rem; margin-top:1rem; }}
.about-point {{ display:flex; gap:.65rem; align-items:flex-start; padding:.8rem; border-radius:13px; background:var(--a-surface-2); color:var(--a-muted); font-size:.9rem; line-height:1.45; }}
.about-point::before {{ content:"✓"; flex:0 0 auto; color:var(--a-success); font-weight:900; }}
.contact-actions {{ display:flex; gap:.6rem; flex-wrap:wrap; margin-top:1rem; }}
.contact-note {{ margin-top:.8rem; padding:.75rem .85rem; border-radius:12px; background:var(--a-primary-soft); color:var(--a-muted); font-size:.82rem; line-height:1.5; }}
@keyframes accountra-rise {{ from {{ opacity:0; transform:translateY(9px); }} to {{ opacity:1; transform:translateY(0); }} }}
@media (prefers-reduced-motion: reduce) {{ .hero, .dashboard-card, .page-hero {{ animation:none!important; }} }}
.coming-soon {{ display:inline-flex; margin-top:1.1rem; padding:.55rem .8rem; border:1px solid var(--a-line); border-radius:999px; background:var(--a-surface-2); color:var(--a-muted); font-size:.82rem; font-weight:800; }}
.back-label {{ padding:.75rem 0; color:var(--a-muted); font-size:.92rem; font-weight:650; }}


.feature-grid {{ display:grid; grid-template-columns:repeat(3,1fr); gap:1rem; margin-top:1rem; }}
.feature {{ background:var(--a-surface); border:1px solid var(--a-line); border-radius:18px; padding:1.35rem; box-shadow:0 5px 20px rgba(15,23,42,.035); transition:transform .18s ease, box-shadow .18s ease, border-color .18s ease; min-height:145px; }}
.feature:hover {{ transform:translateY(-3px); box-shadow:var(--a-shadow); border-color:rgba(88,101,242,.35); }}
.feature-num {{ color:var(--a-primary); font-size:.78rem; font-weight:900; letter-spacing:.1em; }}
.feature-title {{ font-size:1.18rem; font-weight:850; margin:.6rem 0 .35rem; color:var(--a-ink); }}
.feature-copy {{ color:var(--a-muted); line-height:1.55; font-size:.94rem; }}

.section-title {{ font-size:1.65rem; font-weight:900; letter-spacing:-.03em; color:var(--a-ink); margin:2.6rem 0 .35rem; }}
.section-copy {{ color:var(--a-muted); font-size:1rem; margin-bottom:1rem; }}
.info-grid {{ display:grid; grid-template-columns:repeat(3,1fr); gap:1rem; }}
.info-card {{ border:1px solid var(--a-line); border-radius:18px; padding:1.25rem; background:var(--a-surface); }}
.info-card strong {{ display:block; font-size:1.1rem; margin-bottom:.4rem; color:var(--a-ink); }}
.info-card span {{ color:var(--a-muted); line-height:1.55; }}

.workspace-head {{ display:flex; align-items:flex-start; justify-content:space-between; gap:1rem; margin-bottom:1.2rem; }}
.workspace-title {{ font-size:2.25rem; font-weight:900; letter-spacing:-.04em; color:var(--a-ink); }}
.workspace-sub {{ color:var(--a-muted); font-size:1rem; margin-top:.25rem; }}

.panel {{ background:var(--a-surface); border:1px solid var(--a-line); border-radius:var(--a-radius); padding:1.45rem; box-shadow:0 6px 25px rgba(15,23,42,.035); }}
.panel-title {{ font-size:1.25rem; font-weight:850; color:var(--a-ink); }}
.panel-copy {{ color:var(--a-muted); margin-top:.35rem; line-height:1.55; }}
.upload-panel {{ min-height:330px; display:flex; flex-direction:column; justify-content:center; }}
.upload-zone {{ border:1.5px dashed rgba(88,101,242,.45); border-radius:18px; padding:2rem; text-align:center; background:linear-gradient(180deg,rgba(88,101,242,.05),transparent); margin-top:1.2rem; }}
.upload-zone-title {{ font-size:1.3rem; font-weight:850; color:var(--a-ink); }}
.upload-zone-copy {{ color:var(--a-muted); margin-top:.35rem; }}
.context-panel {{ position:sticky; top:1rem; }}
.context-label {{ font-size:.82rem; text-transform:uppercase; letter-spacing:.09em; font-weight:900; color:var(--a-primary); margin-bottom:.9rem; }}
.context-panel-heading {{ padding:.15rem 0 .85rem; border-bottom:1px solid var(--a-line); margin-bottom:.9rem; }}
.context-panel-heading .context-label {{ margin-bottom:.25rem; }}
.context-panel-help {{ color:var(--a-muted); font-size:.86rem; line-height:1.5; }}

.kpi-grid {{ display:grid; grid-template-columns:repeat(4,1fr); gap:1rem; margin:1.2rem 0; }}
.kpi {{ background:var(--a-surface); border:1px solid var(--a-line); border-radius:18px; padding:1.2rem 1.3rem; min-height:125px; }}
.kpi-label {{ color:var(--a-muted); font-weight:750; font-size:.92rem; }}
.kpi-value {{ color:var(--a-ink); font-size:2rem; font-weight:900; letter-spacing:-.035em; margin-top:.45rem; }}
.kpi-note {{ color:var(--a-muted); font-size:.8rem; margin-top:.3rem; }}

.nav-strip {{ display:flex; gap:.55rem; flex-wrap:wrap; padding:.75rem; border:1px solid var(--a-line); background:var(--a-surface); border-radius:16px; margin:1rem 0 1.3rem; }}
.nav-help {{ color:var(--a-muted); font-size:.84rem; margin:.35rem .1rem .2rem; }}

.workflow-stepper {{ display:grid; grid-template-columns:repeat(6,minmax(0,1fr)); gap:.45rem; padding:.55rem; margin:-.35rem 0 1.25rem; border:1px solid var(--a-line); border-radius:16px; background:var(--a-surface); box-shadow:0 4px 18px rgba(15,23,42,.025); }}
.workflow-step {{ display:flex; align-items:center; gap:.55rem; min-width:0; padding:.65rem .7rem; border-radius:11px; color:var(--a-muted); font-size:.76rem; font-weight:850; line-height:1.2; }}
.workflow-step-marker {{ display:grid; place-items:center; flex:0 0 1.55rem; width:1.55rem; height:1.55rem; border:1px solid var(--a-line-strong); border-radius:50%; color:var(--a-muted); font-size:.7rem; font-variant-numeric:tabular-nums; }}
.workflow-step-label {{ overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }}
.workflow-step:focus-visible, .stButton button:focus-visible, input:focus-visible, textarea:focus-visible, [role="combobox"]:focus-visible {{ outline:3px solid rgba(88,101,242,.32)!important; outline-offset:2px!important; }}
.workflow-step.is-active {{ background:var(--a-primary-soft); color:var(--a-primary); }}
.workflow-step.is-active .workflow-step-marker {{ border-color:var(--a-primary); background:var(--a-primary); color:#fff; box-shadow:0 0 0 3px rgba(88,101,242,.12); }}
.workflow-step.is-complete {{ color:var(--a-ink); }}
.workflow-step.is-complete .workflow-step-marker {{ border-color:rgba(22,163,74,.28); background:rgba(22,163,74,.09); color:var(--a-success); }}

.phase6-status {{ display:flex; align-items:center; justify-content:space-between; gap:1rem; margin:-.35rem 0 1rem; padding:.85rem 1rem; border:1px solid var(--a-line); border-radius:15px; background:linear-gradient(135deg,var(--a-surface),var(--a-surface-2)); }}
.phase6-status-main {{ min-width:0; }}
.phase6-status-eyebrow {{ color:var(--a-primary); font-size:.68rem; font-weight:900; letter-spacing:.1em; text-transform:uppercase; }}
.phase6-status-title {{ color:var(--a-ink); font-size:.98rem; font-weight:900; margin-top:.16rem; }}
.phase6-status-copy {{ color:var(--a-muted); font-size:.82rem; line-height:1.45; margin-top:.18rem; }}
.phase6-status-meta {{ flex:0 0 auto; display:flex; align-items:center; gap:.45rem; color:var(--a-muted); font-size:.76rem; font-weight:800; white-space:nowrap; }}
.phase6-status-dot {{ width:.45rem; height:.45rem; border-radius:50%; background:var(--a-success); box-shadow:0 0 0 4px rgba(22,163,74,.1); }}
.phase6-empty {{ margin:1rem 0; padding:2rem 1.3rem; border:1px dashed var(--a-line-strong); border-radius:18px; background:var(--a-surface-2); text-align:center; }}
.phase6-empty-icon {{ display:grid; place-items:center; width:2.5rem; height:2.5rem; margin:0 auto .7rem; border-radius:50%; background:var(--a-primary-soft); color:var(--a-primary); font-size:1.1rem; font-weight:900; }}
.phase6-empty-title {{ color:var(--a-ink); font-size:1.08rem; font-weight:900; }}
.phase6-empty-copy {{ max-width:40rem; margin:.35rem auto 0; color:var(--a-muted); line-height:1.55; font-size:.9rem; }}
.phase6-next {{ margin:1rem 0 1.2rem; padding:.85rem 1rem; border-left:3px solid var(--a-primary); border-radius:0 12px 12px 0; background:var(--a-primary-soft); color:var(--a-muted); font-size:.86rem; line-height:1.5; }}
.phase6-next strong {{ color:var(--a-ink); }}
.upload-format-note {{ margin-top:.65rem; color:var(--a-muted); font-size:.78rem; line-height:1.5; }}

.back-row {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:1rem; }}
.back-label {{ color:var(--a-muted); font-weight:700; font-size:.9rem; }}

.metric-big {{ font-size:2.5rem; font-weight:900; letter-spacing:-.045em; color:var(--a-ink); }}
.status-good {{ color:var(--a-success); font-weight:850; }}
.status-warn {{ color:var(--a-warning); font-weight:850; }}
.status-bad {{ color:var(--a-danger); font-weight:850; }}

.ai-card {{ border:1px solid rgba(88,101,242,.22); background:linear-gradient(135deg,rgba(88,101,242,.08),rgba(6,182,212,.06)); border-radius:18px; padding:1.35rem; }}
.ai-title {{ font-size:1.2rem; font-weight:900; color:var(--a-ink); }}
.ai-copy {{ color:var(--a-muted); line-height:1.6; margin-top:.35rem; }}
.review-hero {{ display:flex; align-items:center; justify-content:space-between; gap:1rem; margin:1rem 0; padding:1.15rem 1.25rem; border:1px solid rgba(245,158,11,.24); border-radius:18px; background:linear-gradient(135deg,rgba(245,158,11,.09),rgba(88,101,242,.045)); }}
.review-hero-title {{ color:var(--a-ink); font-weight:900; font-size:1.12rem; }}
.review-hero-copy {{ color:var(--a-muted); margin-top:.3rem; line-height:1.5; font-size:.9rem; }}
.review-count {{ flex:0 0 auto; min-width:4.6rem; padding:.65rem .75rem; border-radius:14px; background:var(--a-surface); border:1px solid rgba(245,158,11,.24); color:var(--a-warning); text-align:center; font-size:1.55rem; font-weight:900; line-height:1; }}
.review-count small {{ display:block; margin-top:.3rem; color:var(--a-muted); font-size:.62rem; font-weight:850; text-transform:uppercase; letter-spacing:.06em; }}
.review-detail {{ margin-top:1rem; padding:1.1rem 1.2rem; border:1px solid var(--a-line); border-radius:16px; background:var(--a-surface); box-shadow:var(--acc-shadow-sm); }}
.review-detail-grid {{ display:grid; grid-template-columns:repeat(3,1fr); gap:.65rem; margin-top:1rem; }}
.review-detail-item {{ padding:.7rem .8rem; border-radius:12px; background:var(--a-surface-2); color:var(--a-muted); font-size:.78rem; }}
.review-detail-item strong {{ display:block; margin-top:.2rem; color:var(--a-ink); font-size:.92rem; }}
.status-badge {{ display:inline-flex; align-items:center; gap:.35rem; padding:.32rem .55rem; border-radius:999px; font-size:.7rem; font-weight:850; }}
.status-badge.good {{ color:var(--a-success); background:rgba(22,163,74,.09); border:1px solid rgba(22,163,74,.2); }}
.status-badge.warn {{ color:var(--a-warning); background:rgba(217,119,6,.09); border:1px solid rgba(217,119,6,.2); }}
.status-badge.neutral {{ color:var(--a-primary); background:var(--a-primary-soft); border:1px solid rgba(88,101,242,.18); }}
.statement-summary {{ display:grid; grid-template-columns:repeat(4,1fr); gap:.75rem; margin:1rem 0 1.2rem; }}
.statement-summary-card {{ padding:.9rem 1rem; border:1px solid var(--a-line); border-radius:14px; background:var(--a-surface); }}
.statement-summary-label {{ color:var(--a-muted); font-size:.7rem; font-weight:850; text-transform:uppercase; letter-spacing:.06em; }}
.statement-summary-value {{ margin-top:.35rem; color:var(--a-ink); font-size:1.25rem; font-weight:900; letter-spacing:-.025em; }}
.validation-hero {{ display:flex; align-items:center; justify-content:space-between; gap:1rem; margin:1rem 0 1.1rem; padding:1.15rem 1.25rem; border:1px solid var(--a-line); border-radius:18px; background:linear-gradient(135deg,var(--a-surface),var(--a-surface-2)); }}
.validation-score {{ display:flex; align-items:baseline; gap:.45rem; color:var(--a-ink); font-size:2.35rem; font-weight:900; letter-spacing:-.05em; }}
.validation-score small {{ color:var(--a-muted); font-size:.8rem; font-weight:750; letter-spacing:0; }}
.validation-copy {{ color:var(--a-muted); line-height:1.5; font-size:.88rem; }}
.check-card {{ padding:.78rem .85rem; border:1px solid var(--a-line); border-radius:12px; background:var(--a-surface); font-size:.86rem; }}
.check-card.pass {{ border-color:rgba(22,163,74,.22); background:rgba(22,163,74,.045); }}
.check-card.review {{ border-color:rgba(217,119,6,.24); background:rgba(217,119,6,.045); }}
.check-card strong {{ display:block; margin-bottom:.22rem; color:var(--a-ink); }}
.check-card span {{ color:var(--a-muted); font-size:.76rem; }}
.export-hero {{ margin:1rem 0 1.15rem; padding:1.2rem 1.3rem; border:1px solid rgba(22,163,74,.2); border-radius:18px; background:linear-gradient(135deg,rgba(22,163,74,.07),rgba(88,101,242,.045)); }}
.export-hero-title {{ display:flex; align-items:center; gap:.5rem; color:var(--a-ink); font-size:1.12rem; font-weight:900; }}
.export-hero-title::before {{ content:"✓"; display:grid; place-items:center; width:1.45rem; height:1.45rem; border-radius:50%; background:var(--a-success); color:#fff; font-size:.8rem; }}
.export-hero-copy {{ margin-top:.35rem; color:var(--a-muted); line-height:1.5; font-size:.88rem; }}
.export-card {{ min-height:112px; padding:1rem 1.05rem; border:1px solid var(--a-line); border-radius:15px; background:var(--a-surface); box-shadow:var(--acc-shadow-sm); }}
.export-card strong {{ display:block; color:var(--a-ink); font-size:1rem; }}
.export-card span {{ display:block; margin-top:.35rem; color:var(--a-muted); font-size:.8rem; line-height:1.45; }}
.creator-footer {{ margin:2.5rem 0 .4rem; padding:1rem 0 .2rem; border-top:1px solid var(--a-line); color:var(--a-muted); text-align:center; font-size:.78rem; letter-spacing:.01em; }}
.creator-footer strong {{ color:var(--a-ink); font-weight:900; }}

.feedback {{ margin-top:2rem; padding:1.35rem; border-radius:18px; background:var(--a-surface-2); border:1px solid var(--a-line); }}
.feedback-title {{ font-size:1.1rem; font-weight:850; color:var(--a-ink); }}

/* Make Streamlit controls readable */
.stTextInput input, .stTextArea textarea, .stNumberInput input, .stDateInput input {{ font-size:1.02rem!important; min-height:48px!important; color:var(--a-ink)!important; background:var(--a-surface-2)!important; border-color:var(--a-line)!important; }}
.stSelectbox div[data-baseweb="select"] {{ min-height:48px!important; }}
.stButton button {{ min-height:48px!important; border-radius:12px!important; font-size:1rem!important; font-weight:800!important; transition:transform .16s ease, box-shadow .16s ease!important; }}
.stButton button:hover {{ transform:translateY(-1px); box-shadow:0 8px 22px rgba(15,23,42,.09)!important; }}
.stDataFrame {{ font-size:1rem!important; }}
[data-testid="stMetricValue"] {{ font-size:2rem!important; font-weight:900!important; }}
[data-testid="stMetricLabel"] {{ font-size:.95rem!important; }}
[data-testid="stMarkdownContainer"], [data-testid="stCaptionContainer"], .stTextInput label, .stTextArea label, .stNumberInput label, .stDateInput label, .stSelectbox label {{ font-size:1.02rem; }}
[data-testid="stCaptionContainer"] {{ font-size:.9rem!important; }}

@media (max-width: 1000px) {{
  .main .block-container {{ padding:1.2rem 1.2rem 3rem!important; }}
  .hero {{ padding:3rem 2rem; }}
  .feature-grid,.info-grid {{ grid-template-columns:1fr; }}
  .kpi-grid {{ grid-template-columns:repeat(2,1fr); }}
  .context-panel {{ position:static; }}
}}
@media (max-width: 640px) {{
  .hero {{ padding:2.2rem 1.2rem; border-radius:20px; }}
  .hero h1 {{ font-size:2.65rem!important; }}
  .hero-copy {{ font-size:1rem; }}
  .hero-proof {{ gap:.4rem .75rem; }}
  .kpi-grid {{ grid-template-columns:1fr; }}
  .workspace-title {{ font-size:1.8rem; }}
  .a-top-meta {{ display:none; }}
  .a-top-actions {{ gap:0; }}
  .a-session-pill {{ font-size:.7rem; padding:.4rem .58rem; }}
  .main .block-container {{ padding:.9rem .75rem 2.5rem!important; }}
  .workflow-stepper {{ display:flex; overflow-x:auto; gap:.35rem; scrollbar-width:none; }}
  .workflow-stepper::-webkit-scrollbar {{ display:none; }}
  .workflow-step {{ flex:0 0 7.6rem; flex-direction:column; justify-content:center; gap:.35rem; padding:.6rem .45rem; text-align:center; font-size:.68rem; }}
  .workflow-step-label {{ max-width:100%; }}
  .dashboard-rail, .about-grid {{ grid-template-columns:1fr; }}
  .dashboard-strip {{ align-items:flex-start; flex-direction:column; }}
  .site-nav {{ justify-content:stretch; }}
  .site-nav-note {{ display:none; }}
  .site-nav .stButton {{ flex:1; }}
  .review-hero, .validation-hero {{ align-items:flex-start; flex-direction:column; }}
  .review-count {{ align-self:flex-start; }}
  .review-detail-grid, .statement-summary {{ grid-template-columns:1fr 1fr; }}
  .phase6-status {{ align-items:flex-start; flex-direction:column; gap:.55rem; }}
  .phase6-status-meta {{ white-space:normal; }}
}}
</style>
""", unsafe_allow_html=True)


def v7_site_nav():
    """Top-right informational navigation for the public product shell."""
    st.markdown("<div class='site-nav'><span class='site-nav-note'>Simple, review-first accounting workflow</span>", unsafe_allow_html=True)
    left, right = st.columns([1, 1])
    with left:
        if st.button("About", key="v7_about_nav", use_container_width=True):
            st.session_state["app_page"] = "about"
            st.rerun()
    with right:
        if st.button("Contact Us", key="v7_contact_nav", use_container_width=True):
            st.session_state["app_page"] = "contact"
            st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)


def v7_topbar():
    st.markdown("""
    <div class='a-topbar'>
      <div class='a-brand'><div class='a-logo'>A</div><div><div class='a-brand-name'>Accountra</div><div class='a-brand-sub'>AI-powered accounting workspace</div></div></div>
      <div class='a-top-actions'>
        <div class='a-session-pill'><span class='a-session-dot'></span>Session workspace</div>
        <div class='a-top-meta'>Financial intelligence &middot; Review before filing</div>
      </div>
    </div>
    """, unsafe_allow_html=True)


def v7_public_page(kind):
    """Render the lightweight public About or Contact page."""
    v7_topbar()
    v7_site_nav()
    if kind == "about":
        st.markdown("<section class='page-hero'><div class='page-eyebrow'>ABOUT ACCOUNTRA</div><div class='page-title'>Accounting clarity, with a human in the loop.</div><div class='page-copy'>Accountra helps turn Trial Balance data into structured, review-ready financial statements without hiding the numbers behind a black box.</div></section>", unsafe_allow_html=True)
        st.markdown("<div class='about-grid'><div class='about-card'><strong>Built for the review before the report</strong><p>Accountra combines deterministic accounting rules, AI-assisted classification, validation controls and export-ready statements in one focused workspace.</p><div class='about-points'><div class='about-point'>Upload a Trial Balance in the format you already have.</div><div class='about-point'>Review uncertain classifications before they reach the statements.</div><div class='about-point'>Validate the numbers, then export Excel and PDF reports.</div></div></div><div class='about-card'><strong>Meet the builder</strong><p>Accountra is built by <strong>Rohan A</strong> as an AI-assisted accounting workflow for clearer, more controlled financial reporting.</p><div class='contact-note'>Your uploaded files stay inside the active session until you reset the workspace. Avoid sharing confidential financial information through public contact channels.</div></div></div>", unsafe_allow_html=True)
    else:
        st.markdown("<section class='page-hero'><div class='page-eyebrow'>CONTACT ACCOUNTRA</div><div class='page-title'>Have feedback or an idea?</div><div class='page-copy'>Tell us what worked, what needs attention, or what would make your accounting workflow better.</div></section>", unsafe_allow_html=True)
        st.markdown("<div class='about-grid'><div class='contact-card'><strong>Send feedback</strong><p>Use the Accountra feedback form to share product feedback, bug reports or feature requests.</p></div><div class='contact-card'><strong>Keep sensitive data private</strong><p>Please do not include Trial Balance files, API keys, passwords or confidential client information in a public form.</p><div class='contact-note'>The Google Form link will be configured through the secure ACCOUNTRA_GOOGLE_FORM_URL setting.</div></div></div>", unsafe_allow_html=True)
        if GOOGLE_FORM_URL:
            st.link_button("Open Contact Form →", GOOGLE_FORM_URL, type="primary", use_container_width=True)
        else:
            st.warning("The contact form link is not configured yet. Add your Google Form URL to ACCOUNTRA_GOOGLE_FORM_URL.")
    st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)
    if st.button("← Back to Accountra", key=f"v7_back_public_{kind}", use_container_width=True):
        st.session_state["app_page"] = "dashboard"
        st.rerun()
    v7_creator_footer()


def v7_back_control():
    """Hierarchical back navigation: inner workspaces never jump straight to Dashboard."""
    current = st.session_state.get("workspace_section", "upload")
    if current == "upload":
        label = "← Back to Dashboard"
        target = "dashboard"
    else:
        label = "← Back"
        previous = {
            "trial_balance": "upload",
            "ai_review": "trial_balance",
            "statements": "ai_review",
            "validation": "statements",
            "reports": "validation",
        }
        target = previous.get(current, "upload")
    left, right = st.columns([1.1, 4.9])
    with left:
        if st.button(label, key=f"v9_back_{current}", use_container_width=True):
            st.session_state["workspace_section"] = target
            if target == "dashboard":
                st.session_state["app_page"] = "dashboard"
            st.rerun()
    with right:
        helper = "Your workspace is saved for this session." if current == "upload" else "Use Back to return to the previous step."
        st.markdown(f"<div class='back-label'>{helper}</div>", unsafe_allow_html=True)


def v7_header(title, subtitle):
    st.markdown(f"<div class='workspace-title'>{title}</div><div class='workspace-sub'>{subtitle}</div>", unsafe_allow_html=True)


def v7_stepper(active):
    """Render workflow state without changing the existing navigation logic."""
    steps=[
        ("upload", "Upload"),
        ("trial_balance", "Trial Balance"),
        ("ai_review", "AI Review"),
        ("statements", "Statements"),
        ("validation", "Validate"),
        ("reports", "Export"),
    ]
    active_index=next((i for i,(key,_) in enumerate(steps) if key==active), 0)
    html=[]
    for index,(key,label) in enumerate(steps):
        state="is-active" if index==active_index else "is-complete" if index<active_index else ""
        marker="✓" if index<active_index else str(index+1)
        current=" aria-current='step'" if index==active_index else ""
        current_label=" (current step)" if index==active_index else ""
        html.append(f"<div class='workflow-step {state}' role='listitem' tabindex='0'{current} aria-label='{label}{current_label}'><span class='workflow-step-marker' aria-hidden='true'>{marker}</span><span class='workflow-step-label'>{label}</span></div>")
    st.markdown(f"<div class='workflow-stepper' role='list' aria-label='Accounting workflow progress'>{''.join(html)}</div>", unsafe_allow_html=True)


def v7_workspace_status(active):
    """Explain the current workflow position and the next user action."""
    stages={
        "upload": ("Upload", "Add a balanced Trial Balance and confirm the report context.", "Source file stays in this session."),
        "trial_balance": ("Trial Balance", "Review the classified account list and filter anything that looks unusual.", "Review before statements."),
        "ai_review": ("AI Review", "Resolve flagged accounts or confirm the current classification before reporting.", "Human review checkpoint."),
        "statements": ("Statements", "Read the Profit & Loss and Balance Sheet with current and comparative columns.", "Numbers remain visible."),
        "validation": ("Validation", "Clear every control check before treating the report package as final.", "Final control layer."),
        "reports": ("Export", "Download the working paper and PDF after reviewing the final status.", "Excel + PDF ready."),
    }
    title,copy,meta=stages.get(active,stages["upload"])
    keys=["upload","trial_balance","ai_review","statements","validation","reports"]
    index=keys.index(active)+1 if active in keys else 1
    st.markdown(f"<div class='phase6-status' role='status'><div class='phase6-status-main'><div class='phase6-status-eyebrow'>Step {index} of 6</div><div class='phase6-status-title'>{title}</div><div class='phase6-status-copy'>{copy}</div></div><div class='phase6-status-meta'><span class='phase6-status-dot' aria-hidden='true'></span>{meta}</div></div>",unsafe_allow_html=True)


def v7_empty_state(title, copy, icon="✓"):
    """Render a consistent, accessible empty state for workflow queues."""
    st.markdown(f"<div class='phase6-empty' role='status'><div class='phase6-empty-icon' aria-hidden='true'>{icon}</div><div class='phase6-empty-title'>{title}</div><div class='phase6-empty-copy'>{copy}</div></div>",unsafe_allow_html=True)


def v7_build_results(df):
    results=[]
    total_rows=max(len(df),1)
    progress=st.progress(0, text="Classifying accounts…")
    for i, row in df.iterrows():
        account=str(row["Account"]).strip()
        debit=float(row["Debit"] or 0)
        credit=float(row["Credit"] or 0)
        result=classify_account(account,debit,credit)
        if result is None:
            try:
                result=classify_account_ai(account,debit,credit)
            except Exception:
                result=None
        if result is None:
            result=make_result("Unknown","NEEDS_REVIEW","NEEDS_REVIEW","Unable to classify the account.",True,0,"Manual review required.")
        classification=result.get("classification","NEEDS_REVIEW")
        statement=result.get("statement","NEEDS_REVIEW")
        nature=result.get("nature","Unknown")
        ambiguous=bool(result.get("ambiguous",True))
        confidence=float(result.get("confidence",0) or 0)
        reason=result.get("reason","") or ""
        missing=result.get("missing_information")
        if classification not in APPROVED_HEADS:
            classification="NEEDS_REVIEW"; statement="NEEDS_REVIEW"; ambiguous=True; confidence=min(confidence,.50); reason="AI returned an unapproved classification."; missing="Manual classification required."
        elif classification in APPROVED_EXPENSE_HEADS | APPROVED_INCOME_HEADS:
            statement="Profit & Loss"
        else:
            statement="Balance Sheet"
        results.append({"Account":account,"Debit":debit,"Credit":credit,"Nature":nature,"Classification":classification,"Statement":statement,"Ambiguous":ambiguous,"Confidence":confidence,"Reason":reason,"Missing Information":missing})
        progress.progress((i+1)/total_rows)
    progress.empty()
    return pd.DataFrame(results)


def v7_apply_overrides(results_df):
    out=results_df.copy()
    for key, override in list(st.session_state.items()):
        if not key.startswith("override_"): continue
        account=key.replace("override_", "", 1)
        mask=out["Account"]==account
        out.loc[mask,"Classification"]=override
        out.loc[mask,"Statement"]="Profit & Loss" if override in (APPROVED_INCOME_HEADS|APPROVED_EXPENSE_HEADS) else "Balance Sheet"
        out.loc[mask,"Ambiguous"]=False; out.loc[mask,"Confidence"]=1.0
    return out


def v7_prepare_data(results_df):
    results_df=v7_apply_overrides(results_df)
    pnl_df=results_df[results_df["Statement"]=="Profit & Loss"].copy()
    revenue=pnl_df[pnl_df["Classification"].isin(APPROVED_INCOME_HEADS)]
    revenue_summary=revenue.groupby("Classification")[["Debit","Credit"]].sum() if len(revenue) else pd.DataFrame(columns=["Debit","Credit"])
    if len(revenue_summary): revenue_summary["Net"]=revenue_summary["Credit"]-revenue_summary["Debit"]
    revenue_ops=float(revenue_summary.loc["Revenue from Operations","Credit"]-revenue_summary.loc["Revenue from Operations","Debit"]) if "Revenue from Operations" in revenue_summary.index else 0.0
    other_income=float(revenue_summary.loc["Other Income","Credit"]-revenue_summary.loc["Other Income","Debit"]) if "Other Income" in revenue_summary.index else 0.0
    total_revenue=float(revenue_summary["Net"].sum()) if len(revenue_summary) else 0.0
    pre_tax_heads=APPROVED_EXPENSE_HEADS-{"Tax Expense"}
    expenses=pnl_df[pnl_df["Classification"].isin(pre_tax_heads)]
    expense_summary=expenses.groupby("Classification")[["Debit","Credit"]].sum() if len(expenses) else pd.DataFrame(columns=["Debit","Credit"])
    if len(expense_summary): expense_summary["Net"]=expense_summary["Debit"]-expense_summary["Credit"]
    total_expenses=float(expense_summary["Net"].sum()) if len(expense_summary) else 0.0
    tax=pnl_df[pnl_df["Classification"]=="Tax Expense"]
    tax_summary=tax[["Debit","Credit"]].sum() if len(tax) else pd.Series({"Debit":0.0,"Credit":0.0})
    tax_expense=max(0.0,float(tax_summary["Debit"]-tax_summary["Credit"]))
    pbt=total_revenue-total_expenses
    profit=pbt-tax_expense

    asset=results_df[results_df["Classification"].isin(APPROVED_ASSET_HEADS)]
    asset_summary=asset.groupby("Classification")[["Debit","Credit"]].sum() if len(asset) else pd.DataFrame(columns=["Debit","Credit"])
    if len(asset_summary): asset_summary["Net"]=asset_summary["Debit"]-asset_summary["Credit"]
    total_assets=float(asset_summary["Net"].sum()) if len(asset_summary) else 0.0
    liab=results_df[results_df["Classification"].isin(APPROVED_LIABILITY_HEADS)]
    liability_summary=liab.groupby("Classification")[["Debit","Credit"]].sum() if len(liab) else pd.DataFrame(columns=["Debit","Credit"])
    if len(liability_summary): liability_summary["Net"]=liability_summary["Credit"]-liability_summary["Debit"]
    total_liabilities=float(liability_summary["Net"].sum()) if len(liability_summary) else 0.0
    equity=results_df[results_df["Classification"].isin(APPROVED_EQUITY_HEADS)]
    equity_summary=equity.groupby("Classification")[["Debit","Credit"]].sum() if len(equity) else pd.DataFrame(columns=["Debit","Credit"])
    if len(equity_summary): equity_summary["Net"]=equity_summary["Credit"]-equity_summary["Debit"]
    total_equity=float(equity_summary["Net"].sum()) if len(equity_summary) else 0.0
    total_el=total_equity+profit+total_liabilities

    def group_amount(summary, heads):
        return sum(float(summary.loc[h,"Net"]) for h in heads if h in summary.index)
    non_current_asset_groups=[("Property, Plant and Equipment",["PPE"]),("Intangible Assets",["Intangible Assets"]),("Capital Work-in-Progress",["Capital Work-in-Progress"]),("Intangible Assets under Development",["Intangible Assets Under Development"]),("Investment Property",["Investment Property"]),("Non-current Investments",["Investments"]),("Other Non-current Assets",["Other Non-current Assets"])]
    current_asset_groups=[("Inventories",["Inventories"]),("Trade Receivables",["Trade Receivables"]),("Cash and Cash Equivalents",["Cash & Cash Equivalents"]),("Other Current Assets",["Other Current Assets"])]
    non_current_liability_groups=[("Long-term Borrowings",["Non-current Borrowings"]),("Other Long-term Liabilities",["Other Non-current Liabilities"])]
    current_liability_groups=[("Short-term Borrowings",["Current Borrowings"]),("Trade Payables",["Trade Payables"]),("Other Current Liabilities",["Other Current Liabilities"]),("Short-term Provisions",["Provisions"])]
    share_capital=group_amount(equity_summary,["Share Capital"])
    other_equity=group_amount(equity_summary,["Other Equity","Capital Account"])
    return {"results_df":results_df,"revenue_summary":revenue_summary,"expense_summary":expense_summary,"total_revenue":total_revenue,"revenue_ops":revenue_ops,"other_income":other_income,"total_expenses":total_expenses,"tax_expense":tax_expense,"pbt":pbt,"profit":profit,"asset_summary":asset_summary,"liability_summary":liability_summary,"equity_summary":equity_summary,"total_assets":total_assets,"total_liabilities":total_liabilities,"total_equity":total_equity,"total_el":total_el,"group_amount":group_amount,"non_current_asset_groups":non_current_asset_groups,"current_asset_groups":current_asset_groups,"non_current_liability_groups":non_current_liability_groups,"current_liability_groups":current_liability_groups,"share_capital":share_capital,"other_equity":other_equity}


def v7_previous_map():
    if not st.session_state.get("comparative_results"): return {}
    cdf=pd.DataFrame(st.session_state["comparative_results"])
    if not len(cdf): return {}
    g=cdf.groupby("Classification")[["Debit","Credit"]].sum(); g["Net"]=g["Credit"]-g["Debit"]
    return g["Net"].to_dict()


def v7_statement_rows(data):
    prev=v7_previous_map(); rev=data["revenue_summary"]; exp=data["expense_summary"]
    def amount(summary,head,income=False):
        if head not in summary.index:return 0.0
        return float(summary.loc[head,"Credit"]-summary.loc[head,"Debit"] if income else summary.loc[head,"Net"])
    def previous(head): return float(prev[head]) if head in prev else None
    pnl=[{"label":"I. Revenue from Operations","kind":"section"},{"label":"Revenue from Operations","note":"1","current":amount(rev,"Revenue from Operations",True),"previous":previous("Revenue from Operations"),"kind":"line indent"},{"label":"II. Other Income","kind":"section"},{"label":"Other Income","note":"2","current":amount(rev,"Other Income",True),"previous":previous("Other Income"),"kind":"line indent"},{"label":"III. Total Income","current":data["total_revenue"],"previous":sum(v for k,v in prev.items() if k in APPROVED_INCOME_HEADS) if prev else None,"kind":"total"},{"label":"IV. Expenses","kind":"section"}]
    order=["Cost of Materials Consumed","Purchases","Changes in Inventories","Employee Benefits Expense","Finance Costs","Depreciation & Amortisation","Other Expenses"]
    n=3
    for h in order:
        a=amount(exp,h)
        if abs(a)>.005:
            pnl.append({"label":h,"note":str(n),"current":a,"previous":previous(h),"kind":"line indent"}); n+=1
    pnl += [{"label":"Total Expenses","current":data["total_expenses"],"previous":sum(v for k,v in prev.items() if k in APPROVED_EXPENSE_HEADS and k!="Tax Expense") if prev else None,"kind":"total"},{"label":"Profit Before Tax","current":data["pbt"],"kind":"subtotal"},{"label":"Tax Expense","note":str(n),"current":data["tax_expense"],"previous":previous("Tax Expense"),"kind":"line indent"},{"label":"Profit for the Period","current":data["profit"],"kind":"grand-total"}]
    bs=[{"label":"I. EQUITY AND LIABILITIES","kind":"section"},{"label":"1. Shareholders' Funds","kind":"subsection"},{"label":"Share Capital","note":"1","current":data["share_capital"],"previous":previous("Share Capital"),"kind":"line indent"},{"label":"Other Equity","note":"2","current":data["other_equity"],"previous":previous("Other Equity") if previous("Other Equity") is not None else previous("Capital Account"),"kind":"line indent"},{"label":"2. Non-current Liabilities","kind":"subsection"}]
    n=3
    for label,heads in data["non_current_liability_groups"]:
        a=data["group_amount"](data["liability_summary"],heads)
        if abs(a)>.005: bs.append({"label":label,"note":str(n),"current":a,"previous":sum(previous(h) or 0 for h in heads) if prev else None,"kind":"line indent"}); n+=1
    bs.append({"label":"3. Current Liabilities","kind":"subsection"})
    for label,heads in data["current_liability_groups"]:
        a=data["group_amount"](data["liability_summary"],heads)
        if abs(a)>.005: bs.append({"label":label,"note":str(n),"current":a,"previous":sum(previous(h) or 0 for h in heads) if prev else None,"kind":"line indent"}); n+=1
    bs += [{"label":"Total Equity and Liabilities","current":data["total_el"],"kind":"grand-total"},{"label":"II. ASSETS","kind":"section"},{"label":"1. Non-current Assets","kind":"subsection"}]
    for label,heads in data["non_current_asset_groups"]:
        a=data["group_amount"](data["asset_summary"],heads)
        if abs(a)>.005: bs.append({"label":label,"note":str(n),"current":a,"previous":sum(previous(h) or 0 for h in heads) if prev else None,"kind":"line indent"}); n+=1
    bs.append({"label":"2. Current Assets","kind":"subsection"})
    for label,heads in data["current_asset_groups"]:
        a=data["group_amount"](data["asset_summary"],heads)
        if abs(a)>.005: bs.append({"label":label,"note":str(n),"current":a,"previous":sum(previous(h) or 0 for h in heads) if prev else None,"kind":"line indent"}); n+=1
    bs.append({"label":"Total Assets","current":data["total_assets"],"kind":"grand-total"})
    return pnl,bs


def v7_render_statement(title, subtitle, rows):
    html=[]
    for r in rows:
        cur="" if r.get("current") is None else indian_currency(r.get("current"))
        prev="—" if r.get("previous") is None else indian_currency(r.get("previous"))
        html.append(f"<tr class='fs-row {r.get('kind','line')}'><td>{r.get('label','')}</td><td class='note'>{r.get('note','')}</td><td class='amount'>{cur}</td><td class='amount'>{prev}</td></tr>")
    st.markdown(f"<div class='panel'><div class='panel-title'>{title}</div><div class='panel-copy'>{subtitle}</div><div style='overflow:auto;margin-top:1rem'><table class='statement-table' style='width:100%;font-size:1.05rem'><thead><tr><th style='text-align:left;padding:.8rem'>Particulars</th><th style='padding:.8rem'>Note</th><th style='text-align:right;padding:.8rem'>Current</th><th style='text-align:right;padding:.8rem'>Previous</th></tr></thead><tbody>{''.join(html)}</tbody></table></div></div>", unsafe_allow_html=True)


def v7_feedback():
    st.markdown("<div class='feedback'><div class='feedback-title'>Feedback</div><div style='color:var(--a-muted);margin:.3rem 0 1rem'>Tell us what worked, what didn't, or what you want next.</div></div>", unsafe_allow_html=True)
    if GOOGLE_FORM_URL:
        st.link_button("Open Feedback Form →", GOOGLE_FORM_URL, type="primary", use_container_width=True)
        st.caption("Responses are collected through the Accountra Google Form.")
        return
    with st.form("v7_feedback_form"):
        ftype=st.selectbox("Type",["Feedback","Bug report"],key="v7_feedback_type")
        msg=st.text_area("Your message",height=110,key="v7_feedback_message")
        email=st.text_input("Email (optional)",key="v7_feedback_email")
        if st.form_submit_button("Send Feedback", type="primary"):
            if msg.strip(): st.success("Thanks — your feedback was recorded for this session.")
            else: st.warning("Please enter a message first.")


def v7_confirmed_upload_panel(source):
    """Show the normalized Trial Balance summary and preserve the analyze boundary."""
    meta = st.session_state.get("uploaded_source_meta", {})
    if meta:
        st.markdown(f"<div class='phase6-next'><strong>Detected:</strong> {meta.get('sheet', 'Uploaded file')} · header row {meta.get('header_row', 'auto')} · {meta.get('mode', 'normalized')}.</div>", unsafe_allow_html=True)
    c1,c2,c3=st.columns(3)
    with c1: st.metric("Accounts",f"{len(source):,}")
    with c2: st.metric("Debit",indian_currency(source["Debit"].sum()))
    with c3: st.metric("Credit",indian_currency(source["Credit"].sum()))
    diff=float(source["Debit"].sum()-source["Credit"].sum())
    if abs(diff)<.01: st.success("Trial Balance is balanced.")
    else: st.error(f"Trial Balance is not balanced. Difference: {indian_currency(abs(diff))}")
    if st.button("Analyze Trial Balance →",type="primary",use_container_width=True,key="v7_analyze"):
        if abs(diff)>=.01:
            st.error("Balance the Trial Balance before continuing.")
        else:
            with st.spinner("Preparing your accounting workspace…"):
                rdf=v7_build_results(source)
                st.session_state["results"]=rdf.to_dict("records")
                st.session_state["prepared"]=True
                st.session_state["workspace_section"]="trial_balance"
                st.rerun()


def v7_creator_footer():
    """Temporary launch attribution; remove this single helper when no longer needed."""
    st.markdown("<div class='creator-footer'>Built by <strong>Rohan A</strong> · Accountra AI-assisted accounting workflow</div>", unsafe_allow_html=True)


def v7_dashboard():
    v7_topbar()
    v7_site_nav()
    st.markdown("""
    <section class='hero'>
      <div class='eyebrow'>AI FINANCIAL WORKSPACE</div>
      <h1>From Trial Balance to <span>decision-ready statements.</span></h1>
      <div class='hero-copy'>A calmer way to classify, review, validate and export financial statements.</div>
    </section>
    """, unsafe_allow_html=True)
    st.markdown("<div class='hero-actions'><div class='hero-actions-copy'>Your numbers stay visible. You stay in control.</div><div class='hero-actions-note'>Private session workspace · Excel and PDF outputs</div></div>",unsafe_allow_html=True)
    a,b=st.columns([1.15,1])
    with a:
        if st.button("Let's Get Started →", type="primary", use_container_width=True, key="v7_start"):
            st.session_state["app_page"]="workspace"; st.session_state["workspace_section"]="upload"; st.rerun()
    with b:
        if st.button("Explore the workflow", use_container_width=True, key="v7_explore"):
            st.session_state["app_page"]="workspace"; st.session_state["workspace_section"]="upload"; st.rerun()
    st.markdown("<div class='hero-proof'><span>Human review before export</span><span>Flexible Trial Balance files</span><span>Excel and PDF outputs</span></div>",unsafe_allow_html=True)
    st.markdown("<div class='dashboard-rail'><div class='dashboard-card'><div class='dashboard-card-num'>01 · CLASSIFY</div><div class='dashboard-card-title'>Understand the file</div><div class='dashboard-card-copy'>Find the accounts, balances and review flags.</div></div><div class='dashboard-card'><div class='dashboard-card-num'>02 · VALIDATE</div><div class='dashboard-card-title'>Check the numbers</div><div class='dashboard-card-copy'>Keep every control visible before reporting.</div></div><div class='dashboard-card'><div class='dashboard-card-num'>03 · EXPORT</div><div class='dashboard-card-title'>Share the result</div><div class='dashboard-card-copy'>Download polished working papers and statements.</div></div></div>",unsafe_allow_html=True)
    st.markdown("<div class='dashboard-strip'><div><strong>Review-first accounting</strong><br><span>AI helps with interpretation. You approve what reaches the report.</span></div><span>Schedule III-style workflow</span></div>",unsafe_allow_html=True)
    st.markdown("<div style='height:1.1rem'></div>",unsafe_allow_html=True)
    st.markdown("<div class='coming-soon'>Dark mode · Coming soon</div>", unsafe_allow_html=True)
    st.caption("Accountra · AI-assisted accounting workflow · Review classifications and statutory disclosures before filing.")
    v7_creator_footer()


def v7_context_panel():
    st.markdown("<div class='context-panel-heading'><div class='context-label'>Report context</div><div class='context-panel-help'>These details flow into statements, notes and exports.</div></div>",unsafe_allow_html=True)
    st.text_input("Company / Entity Name", key="company_name")
    st.text_input("CIN / Registration No.", key="cin")
    st.date_input("Date of filing / reporting", key="reporting_date")
    st.number_input("Movement review threshold (%)",1.0,100.0,step=5.0,key="materiality_threshold")
    fy=f"{st.session_state['reporting_date'].year-1}-{str(st.session_state['reporting_date'].year)[-2:]}" if st.session_state['reporting_date'].month<=3 else f"{st.session_state['reporting_date'].year}-{str(st.session_state['reporting_date'].year+1)[-2:]}"
    st.markdown(f"<div style='margin-top:1rem;padding:.9rem;border-radius:12px;background:var(--a-surface-2);color:var(--a-muted)'><strong style='color:var(--a-ink)'>Reporting year</strong><br>{fy}</div>",unsafe_allow_html=True)


def v7_workspace():
    v7_topbar()
    v7_back_control()
    st.markdown("<div class='workspace-head'><div><div class='workspace-title'>Financial workspace</div><div class='workspace-sub'>Prepare, review and export your accounting statements from one controlled workflow.</div></div></div>",unsafe_allow_html=True)

    if not st.session_state.get("prepared") or st.session_state.get("workspace_section") == "upload":
        v7_stepper("upload")
        v7_workspace_status("upload")
        existing_source = st.session_state.get("uploaded_source_df")
        left,right=st.columns([1.7,1],gap="large")
        with left:
            st.markdown("<div class='panel upload-panel'><div class='panel-title'>Upload your Trial Balance</div><div class='panel-copy'>Bring in an Excel, XLS or CSV export. Accountra will look for the Trial Balance data and normalize the balance columns.</div><div class='upload-zone'><div class='upload-zone-title'>Choose your source file</div><div class='upload-zone-copy'>Your data stays in this session until you reset the workspace.</div><div class='upload-format-note'>No strict template required. Clear account and balance labels help Accountra confirm the right table.</div></div></div>",unsafe_allow_html=True)
            if existing_source is not None:
                st.info(f"A Trial Balance is already loaded ({len(existing_source):,} accounts). Remove it below if you want to upload a different file.")
                v7_confirmed_upload_panel(existing_source)
                if st.button("Remove loaded Trial Balance", key="v9_remove_loaded", use_container_width=True):
                    st.session_state["uploaded_source_df"] = None
                    st.session_state.pop("uploaded_source_meta", None)
                    st.session_state.pop("v7_tb_file_signature", None)
                    st.session_state.pop("v7_tb_confirmed", None)
                    st.session_state["prepared"] = False
                    st.session_state["results"] = []
                    st.session_state["reset_nonce"] += 1
                    st.rerun()
            uploaded=st.file_uploader("Upload Trial Balance",type=["xlsx","xls","csv"],label_visibility="collapsed",key=f"v7_upload_{st.session_state['reset_nonce']}")
            if uploaded and existing_source is None:
                st.success(f"File ready: {uploaded.name}")
                if st.button("Remove file & choose another", key="v9_remove_upload", use_container_width=True):
                    st.session_state["uploaded_source_df"] = None
                    st.session_state.pop("uploaded_source_meta", None)
                    st.session_state.pop("v7_tb_file_signature", None)
                    st.session_state.pop("v7_tb_confirmed", None)
                    st.session_state["prepared"] = False
                    st.session_state["results"] = []
                    st.session_state["reset_nonce"] += 1
                    st.rerun()
                try:
                    detected=v7_extract_trial_balance(uploaded)
                    if detected.get("data") is None:
                        st.warning("Accountra could not confidently find Account and balance data in this workbook.")
                        st.info("Try a cleaner export, or ask the AI interpreter to inspect the workbook and prepare a reviewable Trial Balance.")
                        if st.button("Ask AI to interpret this workbook →",type="primary",use_container_width=True,key="v7_ai_interpret_workbook"):
                            with st.spinner("Reading the workbook and preparing a Trial Balance preview…"):
                                ai_source, clarifications=build_ai_trial_balance(extract_source_text(uploaded))
                            if ai_source is None:
                                st.error(str(clarifications))
                            else:
                                st.session_state["uploaded_source_df"]=ai_source[["Account","Debit","Credit"]].copy()
                                st.session_state["uploaded_source_meta"]={"sheet":"AI interpretation","header_row":"n/a","mode":"AI reconstructed Trial Balance","confidence":None,"clarifications":clarifications}
                                st.session_state["v7_tb_confirmed"]=True
                                st.rerun()
                        st.stop()
                    source=detected["data"]
                    preview=source.head(8).copy()
                    preview["Debit"]=preview["Debit"].map(indian_currency)
                    preview["Credit"]=preview["Credit"].map(indian_currency)
                    st.markdown(f"<div class='phase6-next'><strong>Detected:</strong> {detected['sheet']} · {detected['mode']} · {len(source):,} usable account rows.</div>",unsafe_allow_html=True)
                    st.dataframe(preview,use_container_width=True,hide_index=True,height=290)
                    if detected.get("needs_review"):
                        st.warning("More than one possible table or an unusual layout was detected. Confirm the preview before continuing.")
                    else:
                        st.success("This looks like a Trial Balance. Confirm the preview before classification.")
                    if not st.session_state.get("v7_tb_confirmed"):
                        if st.button("Use detected Trial Balance →",type="primary",use_container_width=True,key="v7_confirm_detected"):
                            st.session_state["uploaded_source_df"]=source
                            st.session_state["uploaded_source_meta"]={"sheet":detected["sheet"],"header_row":detected["header_row"],"mode":detected["mode"],"confidence":detected["confidence"]}
                            st.session_state["v7_tb_confirmed"]=True
                            st.rerun()
                        st.stop()
                    source.columns=(source.columns.astype(str).str.strip().str.lower().str.replace('₹','',regex=False).str.replace('(','',regex=False).str.replace(')','',regex=False).str.strip())
                    source=source.rename(columns={"account":"Account","account name":"Account","ledger":"Account","ledger account":"Account","particulars":"Account","debit":"Debit","debits":"Debit","dr":"Debit","credit":"Credit","credits":"Credit","cr":"Credit"})
                    if not {"Account","Debit","Credit"}.issubset(source.columns): st.error(f"Required columns missing. Detected: {source.columns.tolist()}")
                    else:
                        total_names={"total","grand total","trial balance total","subtotal","total trial balance"}
                        source=source[~source["Account"].astype(str).str.strip().str.lower().isin(total_names)].copy()
                        source["Debit"]=clean_number_series(source["Debit"]); source["Credit"]=clean_number_series(source["Credit"])
                        st.session_state["uploaded_source_df"]=source
                        c1,c2,c3=st.columns(3)
                        with c1: st.metric("Accounts",f"{len(source):,}")
                        with c2: st.metric("Debit",indian_currency(source["Debit"].sum()))
                        with c3: st.metric("Credit",indian_currency(source["Credit"].sum()))
                        diff=float(source["Debit"].sum()-source["Credit"].sum())
                        if abs(diff)<.01: st.success("Trial Balance is balanced.")
                        else: st.error(f"Trial Balance is not balanced. Difference: {indian_currency(abs(diff))}")
                        if st.button("Analyze Trial Balance →",type="primary",use_container_width=True,key="v7_analyze"):
                            if abs(diff)>=.01: st.error("Balance the Trial Balance before continuing.")
                            else:
                                with st.spinner("Preparing your accounting workspace…"):
                                    rdf=v7_build_results(source)
                                    st.session_state["results"]=rdf.to_dict("records")
                                    st.session_state["prepared"]=True
                                    st.session_state["workspace_section"]="trial_balance"
                                    st.rerun()
                except Exception:
                    st.error("Could not read this file. Please check that it is a valid Excel or CSV file with Account, Debit and Credit columns.")
                    st.info("If the file opens normally in Excel, save a fresh copy and upload it again.")
        with right:
            v7_context_panel()
        st.markdown("<div class='section-title'>Before you continue</div><div class='section-copy'>Keep the upload clean and let Accountra handle classification and validation after the file is balanced.</div>",unsafe_allow_html=True)
        v7_feedback()
        v7_creator_footer()
        return

    results_df=v7_apply_overrides(pd.DataFrame(st.session_state["results"]))
    st.session_state["results"]=results_df.to_dict("records")
    data=v7_prepare_data(results_df)
    nav=st.session_state.get("workspace_section","trial_balance")
    buttons=[("trial_balance","Trial Balance"),("ai_review","AI Review"),("statements","Financial Statements"),("validation","Validation"),("reports","Reports & Export")]
    v7_stepper(nav)
    v7_workspace_status(nav)
    st.markdown("<div class='nav-help'>Workspace sections &middot; one view at a time</div>",unsafe_allow_html=True)
    cols=st.columns(len(buttons))
    for col,(key,label) in zip(cols,buttons):
        with col:
            if st.button(label, use_container_width=True, type="primary" if nav==key else "secondary", key=f"v7_nav_{key}"):
                st.session_state["workspace_section"]=key; st.rerun()

    if nav=="trial_balance":
        v7_render_trial_balance(results_df)
    elif nav=="ai_review":
        v7_render_ai_review(results_df)
    elif nav=="statements":
        v7_render_statements(data)
    elif nav=="validation":
        v7_render_validation(results_df,data)
    elif nav=="reports":
        v7_render_reports(data)
    v7_creator_footer()


def v7_render_trial_balance(results_df):
    v7_header("Trial Balance","Search, filter and review the classified accounts before moving to statements.")
    total_debit=float(results_df["Debit"].sum()); total_credit=float(results_df["Credit"].sum())
    review_count=int(results_df["Ambiguous"].fillna(True).sum())
    c1,c2,c3,c4=st.columns(4)
    with c1: st.markdown(f"<div class='kpi'><div class='kpi-label'>Accounts</div><div class='kpi-value'>{len(results_df):,}</div><div class='kpi-note'>Processed</div></div>",unsafe_allow_html=True)
    with c2: st.markdown(f"<div class='kpi'><div class='kpi-label'>Total Debit</div><div class='kpi-value'>{indian_currency(total_debit)}</div></div>",unsafe_allow_html=True)
    with c3: st.markdown(f"<div class='kpi'><div class='kpi-label'>Total Credit</div><div class='kpi-value'>{indian_currency(total_credit)}</div></div>",unsafe_allow_html=True)
    with c4: st.markdown(f"<div class='kpi'><div class='kpi-label'>Status</div><div class='kpi-value' style='font-size:1.35rem'>{'Balanced' if abs(total_debit-total_credit)<.01 else 'Review'}</div></div>",unsafe_allow_html=True)
    review_copy="Review the flagged accounts before generating statements." if review_count else "All accounts have passed the current classification review."
    review_label="Accounts need review" if review_count else "Ready for review"
    st.markdown(f"<div class='review-hero'><div><div class='review-hero-title'>{review_label}</div><div class='review-hero-copy'>{review_copy}</div></div><div class='review-count'>{review_count}<small>flagged</small></div></div>",unsafe_allow_html=True)
    q=st.text_input("Search accounts",placeholder="Search by account name…",key="v7_tb_search")
    f1,f2,f3=st.columns(3)
    with f1: nature=st.selectbox("Nature",["All"]+sorted(results_df["Nature"].dropna().astype(str).unique()),key="v7_tb_nature")
    with f2: cls=st.selectbox("Classification",["All"]+sorted(results_df["Classification"].dropna().astype(str).unique()),key="v7_tb_class")
    with f3: sortby=st.selectbox("Sort by",["Account","Debit","Credit","Confidence"],key="v7_tb_sort")
    view=results_df.copy()
    if q: view=view[view["Account"].astype(str).str.contains(q,case=False,na=False)]
    if nature!="All": view=view[view["Nature"]==nature]
    if cls!="All": view=view[view["Classification"]==cls]
    view=view.sort_values(sortby,ascending=(sortby=="Account"))
    disp=view[["Account","Debit","Credit","Nature","Classification","Statement","Confidence","Ambiguous"]].copy()
    disp["Debit"]=disp["Debit"].map(indian_currency); disp["Credit"]=disp["Credit"].map(indian_currency); disp["Confidence"]=disp["Confidence"].map(lambda x:f"{float(x)*100:.0f}%")
    disp["Review"]=disp.pop("Ambiguous").map(lambda value:"Needs review" if bool(value) else "Ready")
    st.dataframe(disp,use_container_width=True,hide_index=True,height=560)


def v7_render_ai_review(results_df):
    v7_header("AI Review","A focused queue for accounts that deserve human attention before reporting.")
    review=results_df[results_df["Ambiguous"]==True].copy()
    low=results_df[pd.to_numeric(results_df["Confidence"],errors="coerce").fillna(0)<.80].copy()
    c1,c2,c3=st.columns(3)
    with c1: st.metric("Needs review",len(review))
    with c2: st.metric("Below 80% confidence",len(low))
    with c3: st.metric("Approved",int((results_df["Classification"].isin(APPROVED_HEADS)).sum()))
    review_copy="Resolve the highlighted accounts, then continue to statements." if len(review) else "No exceptions are waiting for manual classification."
    st.markdown(f"<div class='review-hero'><div><div class='review-hero-title'>Human review queue</div><div class='review-hero-copy'>{review_copy}</div></div><div class='review-count'>{len(review)}<small>open</small></div></div>",unsafe_allow_html=True)
    st.markdown("<div class='ai-card'><div class='ai-title'>Accountra Copilot</div><div class='ai-copy'>Review the exceptions below. Deterministic classifications are retained, while uncertain accounts remain visible for human confirmation.</div></div>",unsafe_allow_html=True)
    if not len(review):
        v7_empty_state("Review queue is clear", "No accounts currently require manual review. You can continue to Statements, then run Validation before exporting.")
        return
    st.dataframe(review[["Account","Nature","Classification","Confidence","Reason","Missing Information"]].assign(Confidence=lambda x:x["Confidence"].map(lambda v:f"{float(v)*100:.0f}%")),use_container_width=True,hide_index=True,height=380)
    selected=st.selectbox("Account to review",review["Account"].tolist(),key="v7_review_account")
    row=review[review["Account"]==selected].iloc[0]
    confidence=float(row["Confidence"] or 0)
    st.markdown(f"<div class='review-detail'><div class='panel-title'>{selected}</div><div class='panel-copy'>This account is held for human confirmation before it flows into the statements.</div><div class='review-detail-grid'><div class='review-detail-item'>Classification<strong>{row['Classification']}</strong></div><div class='review-detail-item'>Confidence<strong>{confidence*100:.0f}%</strong></div><div class='review-detail-item'>Statement<strong>{row['Statement']}</strong></div></div><div class='panel-copy' style='margin-top:1rem'><strong>Why:</strong> {row['Reason'] or 'No additional reason supplied.'}<br><strong>Missing information:</strong> {row['Missing Information'] or 'None recorded.'}</div></div>",unsafe_allow_html=True)
    opts=list(APPROVED_HEADS)
    new=st.selectbox("Change classification",sorted(opts),index=sorted(opts).index(row["Classification"]) if row["Classification"] in opts else 0,key="v7_override")
    if st.button("Apply classification",type="primary",key="v7_apply_override"):
        st.session_state[f"override_{selected}"]=new; st.session_state["results"]=v7_apply_overrides(results_df).to_dict("records"); st.success("Classification updated."); st.rerun()


def v7_render_statements(data):
    v7_header("Financial Statements","Schedule III-style presentation with clear current and comparative columns.")
    pnl,bs=v7_statement_rows(data)
    bs_difference=data["total_assets"]-data["total_el"]
    comparative_label="Loaded" if st.session_state.get("comparative_results") else "Not supplied"
    comparative_tone="good" if st.session_state.get("comparative_results") else "neutral"
    tally_label="Tally passed" if abs(bs_difference)<.01 else "Needs review"
    tally_tone="good" if abs(bs_difference)<.01 else "warn"
    st.markdown(f"<div class='statement-summary'><div class='statement-summary-card'><div class='statement-summary-label'>Profit for period</div><div class='statement-summary-value'>{indian_currency(data['profit'])}</div></div><div class='statement-summary-card'><div class='statement-summary-label'>Total assets</div><div class='statement-summary-value'>{indian_currency(data['total_assets'])}</div></div><div class='statement-summary-card'><div class='statement-summary-label'>Balance Sheet</div><div class='statement-summary-value'><span class='status-badge {tally_tone}'>{tally_label}</span></div></div><div class='statement-summary-card'><div class='statement-summary-label'>Comparative period</div><div class='statement-summary-value'><span class='status-badge {comparative_tone}'>{comparative_label}</span></div></div></div>",unsafe_allow_html=True)
    v7_render_statement("Statement of Profit & Loss",f"For the period ended {st.session_state['reporting_date'].strftime('%d %B %Y')}",pnl)
    st.markdown("<div style='height:1rem'></div>",unsafe_allow_html=True)
    v7_render_statement("Balance Sheet",f"As at {st.session_state['reporting_date'].strftime('%d %B %Y')}",bs)
    diff=bs_difference
    if abs(diff)<.01: st.success("Balance Sheet Tally passed.")
    else: st.error(f"Balance Sheet Tally failed. Difference: {indian_currency(abs(diff))}")


def v7_render_validation(results_df,data):
    v7_header("Validation Center","The final control layer before you export the financial statements.")
    tb=abs(float(results_df["Debit"].sum()-results_df["Credit"].sum()))<.01
    valid=int((results_df["Classification"].isin(APPROVED_HEADS)).sum())==len(results_df)
    numeric=pd.api.types.is_numeric_dtype(results_df["Debit"]) and pd.api.types.is_numeric_dtype(results_df["Credit"])
    pnl_ok=abs(data["total_revenue"]-data["total_expenses"]-data["tax_expense"]-data["profit"])<.01
    bs_ok=abs(data["total_assets"]-data["total_el"])<.01
    review=int(results_df["Ambiguous"].fillna(True).sum())
    checks=[("Trial Balance balances",tb),("All classifications approved",valid),("Debit/Credit numeric",numeric),("P&L reconciles",pnl_ok),("Balance Sheet tallies",bs_ok),("No manual review pending",review==0)]
    passed=sum(x[1] for x in checks)
    score_tone="good" if passed==len(checks) else "warn"
    score_copy="All control checks passed. Your report package is ready for final review." if passed==len(checks) else "Some controls need attention before the statements should be treated as final."
    st.markdown(f"<div class='validation-hero'><div><div class='validation-score'>{passed}<small>/ {len(checks)} checks passed</small></div><div class='validation-copy'>{score_copy}</div></div><span class='status-badge {score_tone}'>{'Ready to export' if passed==len(checks) else 'Review required'}</span></div>",unsafe_allow_html=True)
    check_html=[]
    for name, ok in checks:
        state="pass" if ok else "review"
        label="PASS" if ok else "REVIEW"
        detail="Control check completed successfully." if ok else "Resolve this item before final export."
        check_html.append(f"<div class='check-card {state}'><strong>{label} · {name}</strong><span>{detail}</span></div>")
    st.markdown(f"<div class='validation-grid'>{''.join(check_html)}</div>",unsafe_allow_html=True)
    if passed==len(checks):
        st.markdown("<div class='phase6-next'><strong>Next:</strong> open Reports &amp; Export to download the Excel working paper and PDF statements.</div>",unsafe_allow_html=True)
    else:
        st.markdown("<div class='phase6-next'><strong>Next:</strong> use the review details above to resolve each highlighted control before exporting.</div>",unsafe_allow_html=True)
    if data["pbt"]<-.01 and data["tax_expense"]>.01: st.warning("PBT is negative while Tax Expense is present. Review the tax treatment before filing.")


def v7_export_statement_rows(rows):
    """Convert V7's rich statement-row dictionaries to the tuple format
    expected by the original Excel/PDF report generators.

    The on-screen statement renderer needs current/previous values and row
    kinds. The legacy export functions intentionally accept simple two-column
    rows. Keeping this conversion at the export boundary preserves both
    interfaces without changing the accounting engine.
    """
    exported = []
    for row in rows:
        label = str(row.get("label", ""))
        kind = str(row.get("kind", ""))
        amount = row.get("current", "")
        if kind in {"section", "subsection"} or amount is None:
            amount = ""
        exported.append((label, amount))
    return exported


def v7_render_reports(data):
    v7_header("Reports & Export","Generate the working paper and financial statements after review.")
    results_df=data["results_df"]
    pnl,bs=v7_statement_rows(data)
    validation_rows=[("Trial Balance balances","PASS" if abs(float(results_df['Debit'].sum()-results_df['Credit'].sum()))<.01 else "REVIEW","Core debit/credit check"),("Balance Sheet tally","PASS" if abs(data['total_assets']-data['total_el'])<.01 else "REVIEW","Assets versus equity and liabilities")]
    notes=[]

    # The screen uses rich dictionaries for current/previous statement values,
    # while the original report generators use simple (label, amount) rows.
    # Convert only at the export boundary so the existing report engine stays
    # unchanged and exports contain actual values instead of dictionary keys.
    export_pnl=v7_export_statement_rows(pnl)
    export_bs=v7_export_statement_rows(bs)
    excel=make_schedule3_excel(company_name=st.session_state["company_name"],cin=st.session_state["cin"],reporting_date=st.session_state["reporting_date"],results_df=results_df,pnl_rows=export_pnl,bs_rows=export_bs,validation_rows=validation_rows,notes_rows=notes)
    pdf=make_schedule3_pdf(company_name=st.session_state["company_name"],cin=st.session_state["cin"],reporting_date=st.session_state["reporting_date"],pnl_rows=export_pnl,bs_rows=export_bs,validation_rows=validation_rows,notes_rows=notes)
    review_count=int(results_df["Ambiguous"].fillna(True).sum())
    st.markdown("<div class='export-hero'><div class='export-hero-title'>Your report package is prepared</div><div class='export-hero-copy'>Download the working paper or presentation PDF after reviewing the classifications and validation controls. Keep the generated files with your engagement documentation.</div></div>",unsafe_allow_html=True)
    st.markdown(f"<div class='info-grid'><div class='export-card'><strong>Excel working paper</strong><span>Structured Schedule III-style workbook with classifications, validation information and review context.</span></div><div class='export-card'><strong>PDF statements</strong><span>Readable Profit &amp; Loss and Balance Sheet presentation for review and sharing.</span></div><div class='export-card'><strong>Review status</strong><span>{'No manual review items remain.' if review_count==0 else f'{review_count} account(s) still flagged for review.'}</span></div></div>",unsafe_allow_html=True)
    safe=re.sub(r"[^A-Za-z0-9_-]+","_",st.session_state["company_name"].strip() or "Accountra").strip("_")
    c1,c2=st.columns(2)
    with c1: st.download_button("Download Excel working paper",excel,f"{safe}_Schedule_III_Working_Paper.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True,key="v7_excel")
    with c2: st.download_button("Download PDF financial statements",pdf,f"{safe}_Financial_Statements.pdf",mime="application/pdf",use_container_width=True,key="v7_pdf")



# =========================================================
# ACCOUNTRA — FORCE LIGHT UI ON ALL DEVICES
# =========================================================
# Keep the product visually consistent even when the user's
# OS/browser prefers dark mode. A future dark theme can be
# introduced intentionally instead of inheriting Streamlit's theme.
st.markdown(
    """
    <style>
    :root, html, body {
        color-scheme: light !important;
    }

    html[data-theme="dark"],
    html[data-theme="light"] {
        color-scheme: light !important;
    }

    .stApp,
    [data-testid="stAppViewContainer"],
    [data-testid="stMain"],
    .main,
    .block-container {
        background: #f5f7fb !important;
        color: #111827 !important;
    }

    /* Prevent Streamlit/BaseWeb controls from inheriting the device's
       dark color scheme on laptops and phones. */
    input,
    textarea,
    button,
    select,
    [role="combobox"],
    [data-baseweb="input"],
    [data-baseweb="select"],
    [data-baseweb="textarea"] {
        color-scheme: light !important;
    }

    .stTextInput input,
    .stTextArea textarea,
    .stNumberInput input,
    .stDateInput input,
    [data-baseweb="input"] input,
    [data-baseweb="textarea"] textarea {
        background: #f8fafc !important;
        color: #111827 !important;
        caret-color: #111827 !important;
    }

    [data-baseweb="select"] > div,
    [data-baseweb="input"] > div,
    [data-baseweb="textarea"] > div {
        background: #f8fafc !important;
        color: #111827 !important;
    }

    [data-baseweb="select"] *,
    [data-baseweb="input"] *,
    [data-baseweb="textarea"] * {
        color: #111827 !important;
    }

    /* Streamlit metric values and common text containers stay readable. */
    [data-testid="stMetricValue"],
    [data-testid="stMetricLabel"],
    [data-testid="stMetricDelta"],
    [data-testid="stCaptionContainer"],
    [data-testid="stMarkdownContainer"] {
        color: #111827;
    }

    /* Keep tables/data grids from inheriting a dark browser palette. */
    [data-testid="stDataFrame"],
    [data-testid="stTable"] {
        color-scheme: light !important;
    }

    /* Mobile readability: give financial figures and controls more room. */
    @media (max-width: 700px) {
        .main .block-container {
            padding: 1rem 1rem 3rem !important;
        }
        .kpi-value { font-size: 2.15rem !important; }
        .kpi-label { font-size: 1rem !important; }
        .workspace-title { font-size: 1.85rem !important; }
        .panel-title { font-size: 1.18rem !important; }
        .panel-copy, .feature-copy, .info-card span { font-size: 1rem !important; }
        .stButton button, .stDownloadButton button { min-height: 48px !important; font-size: 1rem !important; }
        .stTextInput input, .stTextArea textarea, .stNumberInput input, .stDateInput input { font-size: 1rem !important; min-height: 48px !important; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# =========================================================
# ROUTE
# =========================================================

if st.session_state.get("app_page") == "about":
    v7_public_page("about")
elif st.session_state.get("app_page") == "contact":
    v7_public_page("contact")
elif st.session_state.get("app_page") == "dashboard":
    v7_dashboard()
else:
    v7_workspace()
